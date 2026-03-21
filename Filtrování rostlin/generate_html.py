#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel to HTML Bylinky Generator
Načte Excel soubor a vygeneruje kompletní HTML soubor s bylinkami
"""

import sys
import os
import re
import json

# Try to install required packages if missing
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    import openpyxl


def normalize_for_class(text):
    """
    Normalizuje text pro použití v HTML třídě.
    Odstraní diakritiku, převede na malá písmena, nahradí mezery podtržítky.
    """
    if not text:
        return ""
    
    # Převod na string
    text = str(text).strip()
    if not text:
        return ""
    
    # Mapování českých znaků
    char_map = {
        'á': 'a', 'č': 'c', 'ď': 'd', 'é': 'e', 'ě': 'e', 'í': 'i',
        'ň': 'n', 'ó': 'o', 'ř': 'r', 'š': 's', 'ť': 't', 'ú': 'u',
        'ů': 'u', 'ý': 'y', 'ž': 'z',
        'Á': 'a', 'Č': 'c', 'Ď': 'd', 'É': 'e', 'Ě': 'e', 'Í': 'i',
        'Ň': 'n', 'Ó': 'o', 'Ř': 'r', 'Š': 's', 'Ť': 't', 'Ú': 'u',
        'Ů': 'u', 'Ý': 'y', 'Ž': 'z'
    }
    
    # Nahrazení znaků
    normalized = ""
    for char in text:
        normalized += char_map.get(char, char)
    
    # Převod na malá písmena
    normalized = normalized.lower()
    
    # Nahrazení mezer a speciálních znaků podtržítky
    normalized = re.sub(r'[^a-z0-9_]', '_', normalized)
    
    # Odstranění vícenásobných podtržítek
    normalized = re.sub(r'_+', '_', normalized)
    
    # Odstranění úvodních/koncových podtržítek
    normalized = normalized.strip('_')
    
    return normalized


# Soubor s mapou zobrazení tagů (můžete editovat pro kontrolu diakritiky a textů na tlačítkách)
TAG_ZOBRAZENI_FILE = "tag_zobrazeni.json"

# Pořadí a popisky kategorií (sloupců) pro tag_zobrazeni.json – aby bylo zřejmé, odkud tagy pochází
CATEGORY_ORDER = [
    "skupina", "nemoci", "tcm", "ucinky", "cast", "sber", "stanoviste", "barva", "bezpecnost"
]
CATEGORY_LABELS = {
    "skupina": "Skupina (bylina, keř, strom, řasa…)",
    "nemoci": "Nemoci",
    "tcm": "Orgány (TČM)",
    "ucinky": "Účinky",
    "cast": "Použité části rostliny",
    "sber": "Sběr (měsíce, část: měsíc)",
    "stanoviste": "Stanoviště",
    "barva": "Barva květů",
    "bezpecnost": "Bezpečnost",
    "_ostatni": "Ostatní (ze starého souboru nebo mimo výše)",
}


def _flatten_display_map(data):
    """Sjednotí mapu na plochý slovník: pokud je data podle kategorií, sloučí ho."""
    if not data or not isinstance(data, dict):
        return {}
    first_val = next(iter(data.values()), None)
    if isinstance(first_val, dict):
        flat = {}
        for cat, sub in data.items():
            if cat == "_popis_kategorii":
                continue
            if isinstance(sub, dict):
                for k, v in sub.items():
                    if isinstance(v, str):
                        flat[k] = v
            elif isinstance(sub, str):
                flat[cat] = sub
        return flat
    return dict(data)


def get_display_map():
    """
    Načte mapu zobrazení tagů z tag_zobrazeni.json. Pokud soubor neexistuje, vrátí vestavěnou DIAKRITIKA_MAP.
    Podporuje jak plochý formát { "tag": "zobrazení" }, tak členění podle sloupců { "skupina": { "tag": "zobrazení" }, ... }.
    Vždy vrací plochý slovník pro použití v kódu.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(script_dir, TAG_ZOBRAZENI_FILE)
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return _flatten_display_map(data)
        except (json.JSONDecodeError, IOError) as e:
            print(f"Varování: Nepodařilo se načíst {TAG_ZOBRAZENI_FILE}: {e}. Použita vestavěná mapa.")
    return DIAKRITIKA_MAP.copy()


def load_tag_zobrazeni_categorized():
    """
    Načte tag_zobrazeni.json a vrátí strukturu podle kategorií (sloupců).
    Pokud je soubor v plochém formátu, vrátí None a plochou mapu je třeba vzít z get_display_map().
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(script_dir, TAG_ZOBRAZENI_FILE)
    if not os.path.exists(path):
        return None
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except (json.JSONDecodeError, IOError):
        return None
    if not data or not isinstance(data, dict):
        return None
    first_val = next(iter(data.values()), None)
    if isinstance(first_val, dict):
        return data
    return None


# Mapování hodnot bez diakritiky na hodnoty s diakritikou (výchozí; použije se, pokud neexistuje tag_zobrazeni.json)
DIAKRITIKA_MAP = {
        # Měsíce
        'leden': 'leden',
        'unor': 'únor',
        'brezen': 'březen',
        'duben': 'duben',
        'kveten': 'květen',
        'cerven': 'červen',
        'cervenec': 'červenec',
        'srpen': 'srpen',
        'zari': 'září',
        'rijen': 'říjen',
        'listopad': 'listopad',
        'prosinec': 'prosinec',
        'celorocne': 'celoročně',
        
        # Části rostlin
        'koren': 'kořen',
        'kura': 'kůra',
        'list': 'list',
        'kvet': 'květ',
        'plod': 'plod',
        'semena': 'semena',
        'pupeny': 'pupeny',
        'jehlice': 'jehlice',
        'pryskyrice': 'pryskyřice',
        'nať': 'nať',
        'stonky': 'stonky',
        'kvetiny': 'květiny',
        'okvetni_listek': 'okvětní lístek',
        
        # Nemoci a stavy
        'nachlazeni': 'nachlazení',
        'chripka': 'chřipka',
        'kasel': 'kašel',
        'zahleneni': 'zahlenění',
        'zanety': 'záněty',
        'zanety_dychacich_cest': 'záněty dýchacích cest',
        'bolest_v_krku': 'bolest v krku',
        'posileni_imunity': 'posílení imunity',
        'dezinfekce_dychacich_cest': 'dezinfekce dýchacích cest',
        'revmaticke_bolesti': 'revmatické bolesti',
        'bolesti_svalu': 'bolesti svalů',
        'bolesti_kloubu': 'bolesti kloubů',
        'unava': 'únava',
        'vycerpani': 'vyčerpání',
        'kozni_zanety': 'kožní záněty',
        'obklady': 'obklady',
        'hojeni': 'hojení',
        'bronchitida': 'bronchitida',
        'rym': 'rýma',
        'dutiny': 'dutiny',
        'astma': 'astma',
        'svalove_bolesti': 'svalové bolesti',
        'revmaticke_potize': 'revmatické potíže',
        'kloubni_bolesti': 'kloubní bolesti',
        'rany': 'rány',
        'oslabeni_organismu': 'oslabení organismu',
        'popaleniny': 'popáleniny',
        'spaleniny': 'spáleniny',
        'bodnuti_hmyzem': 'bodnutí hmyzem',
        'otoky': 'otoky',
        'zanety_kuze': 'záněty kůže',
        'bradavice': 'bradavice',
        'kuri_oka': 'kuří oka',
        'bolest_ucha': 'bolest ucha',
        'zanety_ucha': 'záněty ucha',
        'horecka': 'horečka',
        'opary': 'opary',
        'poruchy_traveni': 'poruchy trávení',
        'nadymani': 'nadýmání',
        'krece_traviciho_traktu': 'křeče trávicího traktu',
        'nevolnost': 'nevolnost',
        'infekce': 'infekce',
        'zahrivaci_ucinky': 'zahřívací účinky',
        'podpora_krevniho_obehu': 'podpora krevního oběhu',
        'studene_koncetiny': 'studené končetiny',
        'chladne_koncetiny': 'chladné končetiny',
        'chladne koncetiny': 'chladné končetiny',
        'krece': 'křeče',
        'posileni_vitality': 'posílení vitality',
        'posileni vitality': 'posílení vitality',
        'regulace_krevniho_cukru': 'regulace krevního cukru',
        'antioxidant': 'antioxidant',
        'vysoky_krevni_tlak': 'vysoký krevní tlak',
        'vysoky_cholesterol': 'vysoký cholesterol',
        'kozni_problemy': 'kožní problémy',
        'rekonvalescence': 'rekonvalescence',
        'lehce_projimave_ucinky': 'lehce projimavé účinky',
        'lehce projimave ucinky': 'lehce projimavé účinky',
        'zacpa': 'zácpa',
        'cisteni_strev': 'čištění střev',
        'cisteni strev': 'čištění střev',
        'podpora_traveni': 'podpora trávení',
        'podpora traveni': 'podpora trávení',
        'zanety_mocovych_cest': 'záněty močových cest',
        'zanety mocovych cest': 'záněty močových cest',
        'prujem': 'průjem',
        'zanety_traviciho_traktu': 'záněty trávicího traktu',
        'zanety traviciho traktu': 'záněty trávicího traktu',
        
        # Účinky
        'analgeticke': 'analgetické',
        'antisepticke': 'antiseptické',
        'hojive': 'hojivé',
        'protizanetlive': 'protizánětlivé',
        'protizanetlive_ucinky': 'protizánětlivé účinky',
        'protizanetlive ucinky': 'protizánětlivé účinky',
        'antibakterialni_ucinky': 'antibakteriální účinky',
        'antibakterialni ucinky': 'antibakteriální účinky',
        'protivirove_ucinky': 'protivirové účinky',
        'protivirove ucinky': 'protivirové účinky',
        'zahrivaci': 'zahřívací',
        'sedativni': 'sedativní',
        
        # TCM orgány
        'plice': 'plíce',
        'srdce': 'srdce',
        'jatra': 'játra',
        'slezina': 'slezina',
        'ledviny': 'ledviny',
        'zaludek': 'žaludek',
        'streva': 'střeva',
        
        # Stanoviště
        'suche_svahy': 'suché svahy',
        'skalnate_pudy': 'skalnaté půdy',
        'pobrezni_oblasti': 'pobřežní oblasti',
        'pisecne_pudy': 'písčité půdy',
        'suche_lesy': 'suché lesy',
        'hory': 'hory',
        'lesy': 'lesy',
        'chladna_stanoviste': 'chladná stanoviště',
        'vlhke_louky': 'vlhké louky',
        'vlhke_lesy': 'vlhké lesy',
        'vlhka_stanoviste': 'vlhká stanoviště',
        'okraje_vod': 'okraje vod',
        'okraje_cest': 'okraje cest',
        'ruderalni_plochy': 'ruderální plochy',
        'krovi': 'křoví',
        'horske_oblasti': 'horské oblasti',
        'luzni_lesy': 'lužní lesy',
        'listnate_lesy': 'listnaté lesy',
        'brehy_vod': 'břehy vod',
        'louky': 'louky',
        'pestovane_plochy': 'pěstované plochy',
        'pestovane plochy': 'pěstované plochy',
        'sucha_oblasti': 'suché oblasti',
        'sucha oblasti': 'suché oblasti',
        'polopouste': 'polopouště',
        'mokrady': 'mokřady',
        'sucha_stanoviste': 'suchá stanoviště',
        'sucha stanoviste': 'suchá stanoviště',
        'interiery': 'interiéry',
        'zahrady': 'zahrady',
        'pestovane plochy': 'pěstované plochy',
        'sucha_oblasti': 'suché oblasti',
        'sucha oblasti': 'suché oblasti',
        'polopouste': 'polopouště',
        'mokrady': 'mokřady',
        'sucha_stanoviste': 'suchá stanoviště',
        'sucha stanoviste': 'suchá stanoviště',
        
        # Další nemoci a stavy
        'cisteni_krve': 'čištění krve',
        'cisteni krve': 'čištění krve',
        'podpora_zlucniku': 'podpora žlučníku',
        'podpora zlucniku': 'podpora žlučníku',
        'odvodneni': 'odvodnění',
        'zevni_pouziti': 'zevní použití',
        'zevni pouziti': 'zevní použití',
        'detoxikace': 'detoxikace',
        'podpora_jater': 'podpora jater',
        'zanet_spojivek': 'zánět spojivek',
        'poruchy_pigmentace_kuze': 'poruchy pigmentace kůže',
        'gynekologicke_potize': 'gynekologické potíže',
        'hemoroidy': 'hemoroidy',
        'revmatismus': 'revmatismus',
        'vysoka_hladina_cukru_v_krvi': 'vysoká hladina cukru v krvi',
        'hormonalni_nerovnovaha': 'hormonální nerovnováha',
        'poruchy_menstruace': 'poruchy menstruace',
        'bolestiva_menstruace': 'bolestivá menstruace',
        'bolesti_hlavy': 'bolesti hlavy',
        'kvasinkove_infekce': 'kvasinkové infekce',
        'neplodnost': 'neplodnost',
        'menopauza': 'menopauza',
        'klimaktericke_potize': 'klimakterické potíže',
        'bolesti_bricha': 'bolesti břicha',
        'zanet_jater': 'zánět jater',
        'cirhoza_jater': 'cirhóza jater',
        'prevence_rakoviny': 'prevence rakoviny',
        'nemoci_kuze': 'nemoci kůže',
        'dermatoza': 'dermatoza',
        'alopecie': 'alopecie',
        'poruchy_traveni': 'poruchy trávení',
        'kinetoza': 'kinetóza',
        'nechutenstvi': 'nechutenství',
        'podpora_chuti_k_jidlu': 'podpora chuti k jídlu',
        'zahleneni_prudusek_a_plic': 'zahlenění průdušek a plic',
        'ucpany_nos': 'ucpaný nos',
        'ucpane_nosni_dutiny': 'ucpané nosní dutiny',
        'zpozdena_menstruace': 'zpožděná menstruace',
        'angina': 'angina',
        'zanet_mandli': 'zánět mandlí',
        'zanety_dutiny_ustni': 'záněty dutiny ústní',
        'zanet_prudusek': 'zánět průdušek',
        'slaba_imunita': 'slabá imunita',
        'nemoci_ledvin': 'nemoci ledvin',
        'revma': 'revma',
        'anorexie': 'anorexie',
        'srdce': 'srdce',
        'cevy': 'cévy',
        'krevni_obeh': 'krevní oběh',
        'sucha_kuze': 'suchá kůže',
        'lupenka': 'lupénka',
        'akne': 'akné',
        'vlasy': 'vlasy',
        'nehty': 'nehty',
        'problemy_s_koncentraci': 'problémy s koncentrací',
        'travici_potize': 'trávicí potíže',
        'rakovina_jater': 'rakovina jater',
        'rakovina_prostaty': 'rakovina prostaty',
        'paralyza': 'paralýza',
        'slabost_svalu': 'slabost svalů',
        'tres': 'třes',
        'poruchy_koordinace_pohybu': 'poruchy koordinace pohybu',
        'nervova_slabost': 'nervová slabost',
        'svalove_krece': 'svalové křeče',
        'zavrate': 'závratě',
        'vycerpanost': 'vyčerpanost',
        'slabost': 'slabost',
        'poruchy_krevniho_obehu': 'poruchy krevního oběhu',
        'pocit_zimy': 'pocit zimy',
        'nespavost': 'nespavost',
        'nervove_napeti': 'nervové napětí',
        'slabost_po_horecke_nebo_infekci': 'slabost po horečce nebo infekci',
        'rekonvalescence_po_cevnich_a_nervovych_onemocnenich': 'rekonvalescence po cévních a nervových onemocněních',
        'kozni_potize': 'kožní potíže',
        'puchyre': 'puchýře',
        'svedeni': 'svědění',
        'precitlivelost_na_slunce': 'přecitlivělost na slunce',
        'zanetlive_reakce': 'zánětlivé reakce',
        'podrazdenost': 'podrážděnost',
        'citlivost_pokozky': 'citlivost pokožky',
        'rekonvalescence_po_popaleninach': 'rekonvalescence po popáleninách',
        'ekzemy': 'ekzemy',
        
        # Účinky - rozšíření
        'projimave': 'projímavé',
        'mocopudne': 'močopudné',
        'detoxikacni': 'detoxikační',
        'antidiabeticke': 'antidiabetické',
        'antihypertenzni': 'antihypertenzní',
        'imunostimulacni': 'imunostimulační',
        'rekonvalescencni': 'rekonvalescenční',
        'reguluje_menstruaci': 'reguluje menstruaci',
        'podporuje_traveni': 'podporuje trávení',
        'podporuje_funkci_zaludku': 'podporuje funkci žaludku',
        'podporuje_funkci_slinivky': 'podporuje funkci slinivky',
        'podporuje_funkci_zlucniku': 'podporuje funkci žlučníku',
        'hepatoprotektivni': 'hepatoprotektivní',
        'prokrvuje_vnitrni_organy': 'prokrvuje vnitřní orgány',
        'antirevmaticke': 'antirevmatické',
        'antibakterialni': 'antibakteriální',
        'protivirove': 'protivirové',
        'antimykoticke': 'antimykotické',
        'regeneracni': 'regenerační',
        'protinadorove': 'protinádorové',
        'zpomaluje_rust_rakoviny': 'zpomaluje růst rakoviny',
        'chrani_jatra': 'chrání játra',
        'podporuje_uzdraveni_jater': 'podporuje uzdravení jater',
        'podporuje_chut_k_jidlu': 'podporuje chuť k jídlu',
        'vyzivujici': 'výživující',
        'zvlacnujici': 'zvláčňující',
        'zjemnujici': 'zjemnující',
        'hydratacni': 'hydratační',
        'pomaha_regenerovat_pokozku': 'pomáhá regenerovat pokožku',
        'regenerace_kuze': 'regenerace kůže',
        'regenerace kuze': 'regenerace kůže',
        'podpora_pameti': 'podpora paměti',
        'podpora pameti': 'podpora paměti',
        'uceni': 'učení',
        'uzkost': 'úzkost',
        'zlepseni_kognitivnich_funkci': 'zlepšení kognitivních funkcí',
        'zlepseni kognitivnich funkci': 'zlepšení kognitivních funkcí',
        'posileni_mozku': 'posílení mozku',
        'posileni mozku': 'posílení mozku',
        'metabolicky_syndrom': 'metabolický syndrom',
        'metabolicky syndrom': 'metabolický syndrom',
        'zaludecni_potize': 'žaludeční potíže',
        'zaludecni potize': 'žaludeční potíže',
        'cisteni_jater': 'čištění jater',
        'cisteni jater': 'čištění jater',
        'posileni_vlasu': 'posílení vlasů',
        'regenerace_vlasu': 'regenerace vlasů',
        'zdrave_nehty': 'zdravé nehty',
        'snizuje_krevni_tlak': 'snižuje krevní tlak',
        'snizuje_hladinu_cholesterolu': 'snižuje hladinu cholesterolu',
        'chrani_srdce': 'chrání srdce',
        'chrani_cevy': 'chrání cévy',
        'podporuje_funkci_mozku': 'podporuje funkci mozku',
        'omlazujici': 'omlazující',
        'antioxidacni': 'antioxidační',
        'tonizacni': 'tonizační',
        'protirakovinne': 'protirakovinné',
        'tonizuje_jatra': 'tonizuje játra',
        'tonizuje_a_vyzivuje_krev': 'tonizuje a vyživuje krev',
        'snizuje_hladinu_cukru_v_krvi': 'snižuje hladinu cukru v krvi',
        'reguluje_menstruaci': 'reguluje menstruaci',
        
        # TCM orgány - rozšíření
        'tluste_strevo': 'tlusté střevo',
        'krevni_obeh': 'krevní oběh',
        'zlucnik': 'žlučník',
        
        # Části rostlin - rozšíření
        'semeno': 'semeno',
        'poupata': 'poupata',
        'oddenek': 'oddenek',
        'nat': 'nať',
        'kvet_plod': 'květ|plod',
        'koren_semeno_poupata': 'kořen|semeno|poupata',
        'list_oddenek': 'list|oddenek',
        'list_nat': 'list|nať',
        'pupeny_jehlice_pryskyrice': 'pupeny|jehlice|pryskyřice',
        
        # Barvy
        'zluta': 'žlutá',
        'cervena': 'červená',
        'modra': 'modrá',
        'bila': 'bílá',
        'ruzova': 'růžová',
        'fialova': 'fialová',
        'oranzova': 'oranžová',
        'zelena': 'zelená',
        'hneda': 'hnědá',
        
        # Skupiny
        'strom': 'strom',
        'bylina': 'bylina',
        'ker': 'keř',
        'poloker': 'polokeř',
        'rasa': 'řasa',  # řasa (chaluha apod.)
}


def add_diacritics(text):
    """
    Přidá diakritiku k textu bez diakritiky na základě mapování a heuristiky.
    Pokud text už má diakritiku, vrátí ho beze změny.
    """
    if not text:
        return ""
    
    text = str(text).strip()
    if not text:
        return ""
    
    # Pokud text už obsahuje diakritiku, vrať ho beze změny
    if any(char in text for char in 'áčďéěíňóřšťúůýžÁČĎÉĚÍŇÓŘŠŤÚŮÝŽ'):
        return text
    
    # Normalizuj text pro porovnání (převeď na malá písmena, nahraď mezery a pomlčky podtržítky)
    text_normalized = text.lower().replace(' ', '_').replace('-', '_')
    
    display_map = get_display_map()
    # Zkus najít přesné mapování
    if text_normalized in display_map:
        mapped = display_map[text_normalized]
        # Zachovat velká písmena na začátku, pokud je původní text začínal velkým písmenem
        if text and text[0].isupper():
            mapped = mapped[0].upper() + mapped[1:] if mapped else mapped
        return mapped
    
    # Pokud není v mapě, zkus najít podobné hodnoty (s podtržítky nebo bez)
    # Odstraň podtržítka a zkus znovu
    text_no_underscore = text_normalized.replace('_', '')
    if text_no_underscore in display_map:
        mapped = display_map[text_no_underscore]
        # Zachovat velká písmena na začátku
        if text and text[0].isupper():
            mapped = mapped[0].upper() + mapped[1:] if mapped else mapped
        return mapped
    
    # Pokud není v mapě, vrať původní text
    # (lepší než špatná diakritika)
    return text


def parse_multiple_values(value, separator="|"):
    """
    Rozdělí hodnotu na více hodnot podle separátoru.
    Podporuje separátory " | " a "|"
    """
    # Převod na string a odstranění None
    if value is None:
        return []
    
    if not value:
        return []
    
    value = str(value).strip()
    if not value:
        return []
    
    # Zkus různé separátory v pořadí podle priority
    if " | " in value:
        # Separátor s mezerami (nejčastější)
        values = [v.strip() for v in value.split(" | ")]
    elif "|" in value:
        # Separátor bez mezer
        values = [v.strip() for v in value.split("|")]
    elif ";" in value:
        # Středník jako separátor
        values = [v.strip() for v in value.split(";")]
    elif "," in value:
        # Čárka jako separátor
        values = [v.strip() for v in value.split(",")]
    else:
        # Žádný separátor - jedna hodnota
        values = [value.strip()]
    
    # Odstranit prázdné hodnoty, None a NaN
    return [v for v in values if v and v != "None" and v.lower() != "nan"]


def parse_sber(value):
    """
    Parsuje hodnotu sber, která může mít formát:
    "koren:rijen | listopad" -> ["sber_koren_rijen", "sber_listopad"]
    """
    if not value:
        return []
    
    classes = []
    values = parse_multiple_values(value, "|")
    
    for val in values:
        val = val.strip()
        if ":" in val:
            # Formát "cast:mesic" -> "sber_cast_mesic"
            parts = val.split(":", 1)
            cast = normalize_for_class(parts[0])
            mesic = normalize_for_class(parts[1])
            if cast and mesic:
                classes.append(f"sber_{cast}_{mesic}")
        else:
            # Pouze měsíc -> "sber_mesic"
            mesic = normalize_for_class(val)
            if mesic:
                classes.append(f"sber_{mesic}")
    
    return classes


def map_bezpecnost(value):
    """
    Mapuje číselnou hodnotu bezpečnosti na text.
    1 -> "Bezpečná"
    2 -> "Opatrně"
    3 -> "Nebezpečná" (pro případné rozšíření)
    """
    if not value:
        return ""
    
    value = str(value).strip()
    if value == "1":
        return "Bezpečná"
    elif value == "2":
        return "Opatrně"
    elif value == "3":
        return "Nebezpečná"
    else:
        # Pokud už je text, použij ho
        return value


def generate_bylinka_classes(row_data):
    """
    Vygeneruje seznam HTML tříd pro bylinku z dat řádku.
    Prázdné kategorie se zobrazí jako prefix_ (např. tcm_, cast_, sber_)
    """
    classes = ["bylinka"]
    
    # Skupina (může obsahovat více tagů oddělených | nebo , – každý tag = jedno tlačítko)
    if row_data.get('skupina'):
        skupiny = parse_multiple_values(row_data['skupina'], "|")
        for skup in skupiny:
            skup_norm = normalize_for_class(skup)
            if skup_norm:
                classes.append(f"skupina_{skup_norm}")
    
    # Nemoci (oddělené " | ")
    if row_data.get('nemoci'):
        nemoci = parse_multiple_values(row_data['nemoci'], "|")
        for nemoc in nemoci:
            nemoc_norm = normalize_for_class(nemoc)
            if nemoc_norm:
                classes.append(f"nemoci_{nemoc_norm}")
    
    # TCM orgány (oddělené "|") - vždy přidat, i když prázdné
    if row_data.get('tcm_organy'):
        tcm_organs = parse_multiple_values(row_data['tcm_organy'], "|")
        if tcm_organs:
            for organ in tcm_organs:
                organ_norm = normalize_for_class(organ)
                if organ_norm:
                    classes.append(f"tcm_{organ_norm}")
        else:
            classes.append("tcm_")
    else:
        classes.append("tcm_")
    
    # Účinky (oddělené "|")
    if row_data.get('ucinky'):
        ucinky = parse_multiple_values(row_data['ucinky'], "|")
        for ucinek in ucinky:
            ucinek_norm = normalize_for_class(ucinek)
            if ucinek_norm:
                classes.append(f"ucinky_{ucinek_norm}")
    
    # Část rostliny (oddělené "|") - vždy přidat, i když prázdné
    if row_data.get('cast_rostliny'):
        cast_parts = parse_multiple_values(row_data['cast_rostliny'], "|")
        if cast_parts:
            for cast in cast_parts:
                cast_norm = normalize_for_class(cast)
                if cast_norm:
                    classes.append(f"cast_{cast_norm}")
            # Pokud žádná část nebyla přidána, přidej prázdnou třídu
            if not any(c.startswith("cast_") and c != "cast_" for c in classes):
                classes.append("cast_")
        else:
            classes.append("cast_")
    else:
        classes.append("cast_")
    
    # Sběr (speciální formát) - vždy přidat, i když prázdné
    if row_data.get('sber'):
        sber_classes = parse_sber(row_data['sber'])
        if sber_classes:
            classes.extend(sber_classes)
        else:
            classes.append("sber_")
    else:
        classes.append("sber_")
    
    # Stanoviště (oddělené "|")
    if row_data.get('stanoviste'):
        stanoviste_list = parse_multiple_values(row_data['stanoviste'], "|")
        for stanoviste in stanoviste_list:
            stanoviste_norm = normalize_for_class(stanoviste)
            if stanoviste_norm:
                classes.append(f"stanoviste_{stanoviste_norm}")
    
    # Barva květu (oddělené "|")
    if row_data.get('barva_kvetu'):
        barvy = parse_multiple_values(row_data['barva_kvetu'], "|")
        for barva in barvy:
            barva_norm = normalize_for_class(barva)
            if barva_norm:
                classes.append(f"barva_{barva_norm}")
    
    # Bezpečnost
    if row_data.get('bezpecnost'):
        bezpecnost = map_bezpecnost(row_data['bezpecnost'])
        if bezpecnost:
            classes.append(f"bezpecnost_{bezpecnost}")
    
    return classes


def generate_original_values(row_data):
    """
    Vygeneruje slovník s originálními hodnotami s diakritikou pro každou kategorii.
    Automaticky přidá diakritiku k hodnotám bez diakritiky.
    Tyto hodnoty budou uloženy do data atributů pro zobrazení na webu.
    """
    original_values = {}
    
    # Pomocná funkce pro přidání diakritiky k hodnotě nebo seznamu hodnot
    def add_diacritics_to_value(value):
        if not value:
            return value
        value_str = str(value).strip()
        return add_diacritics(value_str)
    
    def add_diacritics_to_list(values):
        if not values:
            return []
        return [add_diacritics_to_value(v) for v in values]
    
    # Skupina – může být více hodnot (bylina, poloker → dvě tlačítka)
    if row_data.get('skupina'):
        skupiny = parse_multiple_values(row_data['skupina'], "|")
        if skupiny:
            skupiny_with_diacritics = add_diacritics_to_list(skupiny)
            original_values['skupina'] = "|".join(skupiny_with_diacritics)
    
    # Nemoci - seznam hodnot oddělených "|"
    if row_data.get('nemoci'):
        nemoci = parse_multiple_values(row_data['nemoci'], "|")
        if nemoci:
            nemoci_with_diacritics = add_diacritics_to_list(nemoci)
            original_values['nemoci'] = "|".join(nemoci_with_diacritics)
    
    # TCM orgány - seznam hodnot oddělených "|"
    if row_data.get('tcm_organy'):
        tcm_organs = parse_multiple_values(row_data['tcm_organy'], "|")
        if tcm_organs:
            tcm_with_diacritics = add_diacritics_to_list(tcm_organs)
            original_values['tcm'] = "|".join(tcm_with_diacritics)
    
    # Účinky - seznam hodnot oddělených "|"
    if row_data.get('ucinky'):
        ucinky = parse_multiple_values(row_data['ucinky'], "|")
        if ucinky:
            ucinky_with_diacritics = add_diacritics_to_list(ucinky)
            original_values['ucinky'] = "|".join(ucinky_with_diacritics)
    
    # Část rostliny - seznam hodnot oddělených "|"
    if row_data.get('cast_rostliny'):
        cast_parts = parse_multiple_values(row_data['cast_rostliny'], "|")
        if cast_parts:
            cast_with_diacritics = add_diacritics_to_list(cast_parts)
            original_values['cast'] = "|".join(cast_with_diacritics)
    
    # Sběr - originální hodnota (speciální formát, může obsahovat "cast:mesic | mesic")
    if row_data.get('sber'):
        sber_value = str(row_data['sber']).strip()
        # Sběr může mít formát "cast:mesic | mesic", takže potřebujeme speciální zpracování
        sber_parts = parse_multiple_values(sber_value, "|")
        sber_with_diacritics = []
        for part in sber_parts:
            part = part.strip()
            if ":" in part:
                # Formát "cast:mesic"
                cast_mesic = part.split(":", 1)
                cast_part = add_diacritics_to_value(cast_mesic[0])
                mesic_part = add_diacritics_to_value(cast_mesic[1]) if len(cast_mesic) > 1 else ""
                if mesic_part:
                    sber_with_diacritics.append(f"{cast_part}:{mesic_part}")
                else:
                    sber_with_diacritics.append(cast_part)
            else:
                # Pouze měsíc nebo jiná hodnota
                sber_with_diacritics.append(add_diacritics_to_value(part))
        if sber_with_diacritics:
            original_values['sber'] = " | ".join(sber_with_diacritics)
    
    # Stanoviště - seznam hodnot oddělených "|"
    if row_data.get('stanoviste'):
        stanoviste_list = parse_multiple_values(row_data['stanoviste'], "|")
        if stanoviste_list:
            stanoviste_with_diacritics = add_diacritics_to_list(stanoviste_list)
            original_values['stanoviste'] = "|".join(stanoviste_with_diacritics)
    
    # Barva květu - seznam hodnot oddělených "|"
    if row_data.get('barva_kvetu'):
        barvy = parse_multiple_values(row_data['barva_kvetu'], "|")
        if barvy:
            barvy_with_diacritics = add_diacritics_to_list(barvy)
            original_values['barva'] = "|".join(barvy_with_diacritics)
    
    # Bezpečnost - už mapovaná hodnota (má diakritiku)
    if row_data.get('bezpecnost'):
        bezpecnost = map_bezpecnost(row_data['bezpecnost'])
        if bezpecnost:
            original_values['bezpecnost'] = bezpecnost
    
    return original_values


def collect_all_used_tags_from_data(data):
    """
    Projde všechna zobrazená data (řádky s zobrazeni='on') a vrátí:
    - used_flat: plochý slovník norm -> zobrazení (pro zpětnou kompatibilitu)
    - used_by_category: slovník kategorie -> { norm -> zobrazení }, aby šlo ukládat tag_zobrazeni podle sloupců
    """
    used = {}
    by_category = {cat: {} for cat in CATEGORY_ORDER}
    for row_data in data:
        zobrazeni = str(row_data.get('zobrazeni', '')).strip().lower()
        if zobrazeni and zobrazeni != 'on':
            continue
        # Mapování klíče v original_values na název kategorie (tcm_organy -> tcm, cast_rostliny -> cast, barva_kvetu -> barva)
        orig_key_to_cat = {
            'skupina': 'skupina', 'nemoci': 'nemoci', 'tcm': 'tcm', 'ucinky': 'ucinky',
            'cast': 'cast', 'sber': 'sber', 'stanoviste': 'stanoviste', 'barva': 'barva', 'bezpecnost': 'bezpecnost'
        }
        orig = generate_original_values(row_data)
        for key, val_str in orig.items():
            if not val_str:
                continue
            cat = orig_key_to_cat.get(key, key)
            if cat not in by_category:
                by_category[cat] = {}
            parts = [p.strip() for p in re.split(r'\s*\|\s*', str(val_str)) if p.strip()]
            for part in parts:
                norm = normalize_for_class(part)
                if norm:
                    used[norm] = part
                    by_category[cat][norm] = part
        # Z tříd doplníme tagy, které v original_values nebyly (např. složené hodnoty sber)
        classes = generate_bylinka_classes(row_data)
        for cls in classes:
            if cls == 'bylinka' or '_' not in cls:
                continue
            cat = cls.split('_', 1)[0]
            suffix = cls[cls.index('_') + 1:]
            if not suffix:
                continue
            if cat not in by_category:
                by_category[cat] = {}
            if suffix not in used:
                disp = add_diacritics(suffix) or suffix
                used[suffix] = disp
                by_category[cat][suffix] = disp
    return used, by_category


def generate_bylinka_html(row_data):
    """
    Vygeneruje HTML kód pro jednu bylinku.
    Ukládá originální hodnoty s diakritikou do data atributů pro zobrazení na webu.
    """
    classes = generate_bylinka_classes(row_data)
    class_str = " ".join(classes)
    
    nazev_cz = row_data.get('nazev_cz', '').strip()
    nazev_lat = row_data.get('nazev_lat', '').strip()
    url = row_data.get('url', '').strip()
    
    # Získání originálních hodnot s diakritikou
    original_values = generate_original_values(row_data)
    
    html = f'<div class="{class_str}"'
    if url:
        html += f' data-url="{url}"'
    
    # Přidání data atributů s originálními hodnotami
    for key, value in original_values.items():
        if value:
            # Escapování uvozovek v hodnotě
            escaped_value = value.replace('"', '&quot;')
            html += f' data-original-{key}="{escaped_value}"'
    
    html += '>\n'
    
    if nazev_cz:
        html += f'  <div class="cz">{nazev_cz}</div>\n'
    
    if nazev_lat:
        html += f'  <div class="lat">{nazev_lat}</div>\n'
    
    html += '</div>\n\n'
    
    return html


def read_excel_file(filename):
    """
    Načte Excel soubor a vrátí seznam slovníků s daty řádků.
    """
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Načtení hlaviček
    headers = []
    for cell in ws[1]:
        headers.append(cell.value if cell.value else "")
    
    # Mapování hlaviček na standardní názvy (B=česky, C=latinsky v Finální.xlsx)
    header_map = {
        'id': 'id',
        'česky': 'nazev_cz',
        'latinsky': 'nazev_lat',
        'nazev_cz': 'nazev_cz',
        'nazev_lat': 'nazev_lat',
        'url': 'url',
        'skupina': 'skupina',
        'nemoci': 'nemoci',
        'tcm_organy': 'tcm_organy',
        'ucinky': 'ucinky',
        'Cast-rostliny': 'cast_rostliny',
        'sber': 'sber',
        'stanoviste': 'stanoviste',
        'barva kvetu': 'barva_kvetu',
        'Bezpečnost': 'bezpecnost',
        'zobrazení': 'zobrazeni',
        'zobrazeni': 'zobrazeni'
    }
    
    # Normalizace hlaviček - flexibilnější mapování
    normalized_headers = []
    for header in headers:
        if not header:
            normalized_headers.append("")
            continue
            
        header_str = str(header).strip()
        header_lower = header_str.lower().strip()
        
        # Najdi mapování - zkus přesné shody i podobné
        mapped = None
        
        # Nejdřív zkus přesné mapování
        for key, value in header_map.items():
            if key.lower() == header_lower:
                mapped = value
                break
        
        # Pokud ne, zkus najít podobné (obsahuje klíčové slovo)
        if not mapped:
            # Odstraň diakritiku a speciální znaky pro porovnání
            header_normalized = header_lower.replace(" ", "_").replace("-", "_")
            for key, value in header_map.items():
                key_normalized = key.lower().replace(" ", "_").replace("-", "_")
                if key_normalized in header_normalized or header_normalized in key_normalized:
                    mapped = value
                    break
        
        # Pokud stále ne, použij normalizovaný název
        normalized_headers.append(mapped if mapped else header_lower.replace(" ", "_").replace("-", "_"))
    
    # Indexy sloupců: A=0, B=1, C=2 — názvy rostlin vždy ze sloupců B a C
    COL_CESKY = 1   # sloupec B
    COL_LATINSKY = 2  # sloupec C

    # Načtení dat
    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {}
        for col_idx, cell_value in enumerate(row):
            if col_idx < len(normalized_headers):
                key = normalized_headers[col_idx]
                if key:
                    # Zpracuj hodnotu - převeď None na prázdný string, zachovej ostatní
                    if cell_value is None:
                        row_data[key] = ""
                    else:
                        # Převeď na string a očisti
                        value_str = str(cell_value).strip()
                        # Odstraň "None" a "nan" stringy
                        if value_str.lower() in ["none", "nan", ""]:
                            row_data[key] = ""
                        else:
                            row_data[key] = value_str

        # Český a latinský název vždy ze sloupců B a C (nezávisle na hlavičkách)
        def _clean_cell(val):
            if val is None:
                return ""
            s = str(val).strip()
            if s.lower() in ("none", "nan", ""):
                return ""
            return s
        if len(row) > COL_CESKY:
            row_data['nazev_cz'] = _clean_cell(row[COL_CESKY])
        if len(row) > COL_LATINSKY:
            row_data['nazev_lat'] = _clean_cell(row[COL_LATINSKY])
        
        # Přeskoč prázdné řádky
        if not row_data.get('nazev_cz') and not row_data.get('nazev_lat'):
            continue
        
        data.append(row_data)
    
    return data


def read_html_template(filename):
    """
    Načte HTML šablonu a vrátí přední část (před bylinkami) a zadní část (po bylinkách).
    """
    with open(filename, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Najdi začátek sekce s bylinkami
    start_marker = '<div id="byliny">'
    start_idx = content.find(start_marker)
    
    if start_idx == -1:
        raise ValueError("Nepodařilo se najít začátek sekce s bylinkami (<div id=\"byliny\">)")
    
    # Najdi konec sekce s bylinkami - hledáme </div> následovaný prázdnými řádky a <script>
    # Použijeme regex pro flexibilní hledání (toleruje různý počet newlines)
    # Najdeme </div> následovaný jednou nebo více prázdnými řádky a pak <script>
    end_pattern = re.compile(r'</div>\s*\n\s*\n\s*<script>', re.MULTILINE)
    end_match = end_pattern.search(content, start_idx)
    
    if end_match is None:
        # Zkusíme jednodušší variantu - najdeme </div> následovaný <script> (s možnými mezerami/newlines mezi nimi)
        end_pattern2 = re.compile(r'</div>\s+<script>', re.MULTILINE)
        end_match = end_pattern2.search(content, start_idx)
        if end_match is None:
            raise ValueError("Nepodařilo se najít konec sekce s bylinkami (</div> následovaný <script>)")
    
    end_idx = end_match.start()
    
    # Přední část (head, styles, filters, začátek byliny)
    front_part = content[:start_idx + len(start_marker)] + '\n\n'
    
    # Zadní část (konec byliny, scripts, body/html)
    # Začneme od </div> (ne od <script>), aby se zachovala struktura
    back_part = '\n\n' + content[end_idx:]
    
    return front_part, back_part


def generate_html(excel_file, template_file, output_file):
    """
    Hlavní funkce pro generování HTML z Excel souboru.
    """
    print(f"Načítám Excel soubor: {excel_file}")
    data = read_excel_file(excel_file)
    print(f"Načteno {len(data)} bylinek")
    
    print(f"Načítám HTML šablonu: {template_file}")
    front_part, back_part = read_html_template(template_file)
    
    print("Generuji HTML kód bylinek...")
    bylinky_html = ""
    skipped_count = 0
    for i, row_data in enumerate(data, 1):
        # Kontrola zobrazení - zobrazit pouze pokud je "on"
        zobrazeni = str(row_data.get('zobrazeni', '')).strip().lower()
        if zobrazeni and zobrazeni != 'on':
            skipped_count += 1
            continue
        
        bylinky_html += generate_bylinka_html(row_data)
        if i % 50 == 0:
            print(f"  Zpracováno {i}/{len(data)} bylinek...")
    
    print(f"Sestavuji kompletní HTML a ukládám do: {output_file}")
    complete_html = front_part + bylinky_html + back_part
    
    # Vložit mapu diakritiky pro zobrazení tagů v prohlížeči (z tag_zobrazeni.json nebo výchozí)
    display_map = get_display_map()
    diakritika_js = 'window.DIAKRITIKA_FALLBACK = ' + json.dumps(display_map, ensure_ascii=False) + ';'
    complete_html = complete_html.replace('// DIAKRITIKA_FALLBACK_PLACEHOLDER', diakritika_js)
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(complete_html)
    
    # Aktualizovat tag_zobrazeni.json – doplnit naprosto všechny použité tagy, členěné podle sloupců
    used_flat, used_by_category = collect_all_used_tags_from_data(data)
    existing_flat = get_display_map()
    # Sestavit výstup podle kategorií (sloupců), aby bylo v souboru zřejmé, odkud tagy pochází
    full_categorized = {"_popis_kategorii": CATEGORY_LABELS}
    placed = set()
    for cat in CATEGORY_ORDER:
        full_categorized[cat] = {}
        for norm, display in used_by_category.get(cat, {}).items():
            placed.add(norm)
            # Tag už v JSON je → ponechat stávající text (vaše úpravy). Nový tag → vzít z Excelu (display).
            full_categorized[cat][norm] = existing_flat.get(norm, display)
        full_categorized[cat] = dict(sorted(full_categorized[cat].items(), key=lambda x: x[0].lower()))
    # Tagy ze starého souboru nebo z dat, které nejsou v žádné standardní kategorii
    for norm in used_flat:
        if norm not in placed:
            placed.add(norm)
            if "_ostatni" not in full_categorized:
                full_categorized["_ostatni"] = {}
            # Stejná logika: existující tag z JSON ponechat, nový z Excelu
            full_categorized["_ostatni"][norm] = existing_flat.get(norm, used_flat[norm])
    # Tagy, které jsou jen v JSON a v Excelu se nevyskytují, se do JSON již nepřidávají (odstraní se).
    if "_ostatni" in full_categorized:
        full_categorized["_ostatni"] = dict(sorted(full_categorized["_ostatni"].items(), key=lambda x: x[0].lower()))
    script_dir = os.path.dirname(os.path.abspath(output_file))
    tag_path = os.path.join(script_dir, TAG_ZOBRAZENI_FILE)
    with open(tag_path, 'w', encoding='utf-8') as f:
        json.dump(full_categorized, f, ensure_ascii=False, indent=2)
    # Všechna podtržítka v zobrazeních (oranžový text) nahradit mezerou
    n_underscores = replace_display_substring_everywhere("_", " ", tag_path)
    if n_underscores:
        print(f"V zobrazeních nahrazeno {n_underscores}× _ za mezeru.")
    total_tags = sum(len(v) for k, v in full_categorized.items() if isinstance(v, dict) and k != "_popis_kategorii")
    print(f"Aktualizováno {total_tags} tagů v {TAG_ZOBRAZENI_FILE} (členěno podle sloupců: Skupina, Nemoci, …).")
    
    displayed_count = len(data) - skipped_count
    print(f"Hotovo! Vygenerováno {displayed_count} bylinek do souboru {output_file}")
    if skipped_count > 0:
        print(f"  (Přeskočeno {skipped_count} bylinek se zobrazením 'off')")


def replace_display_value_everywhere(old_value, new_value, tag_path=None):
    """
    V tag_zobrazeni.json nahradí u všech položek zobrazení (hodnotu), která se rovná old_value, za new_value.
    Funguje pro plochý i členěný (podle kategorií) formát souboru.
    Vrací počet provedených náhrad.
    """
    if tag_path is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        tag_path = os.path.join(script_dir, TAG_ZOBRAZENI_FILE)
    if not os.path.exists(tag_path):
        print(f"Soubor {tag_path} neexistuje.")
        return 0
    with open(tag_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    count = 0

    def replace_in_obj(obj):
        nonlocal count
        if isinstance(obj, dict):
            for k, v in list(obj.items()):
                if k == "_popis_kategorii":
                    continue
                if isinstance(v, str) and v == old_value:
                    obj[k] = new_value
                    count += 1
                else:
                    replace_in_obj(v)
        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                if isinstance(item, str) and item == old_value:
                    obj[i] = new_value
                    count += 1
                else:
                    replace_in_obj(item)

    replace_in_obj(data)
    with open(tag_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return count


def replace_display_substring_everywhere(old_substring, new_substring, tag_path=None):
    """
    V tag_zobrazeni.json nahradí v každém zobrazení (hodnotě) všude výskyt podřetězce old_substring za new_substring.
    Např. "cisti" -> "čistí" změní "cisti ledviny" na "čistí ledviny", "cisti" na "čistí" atd.
    Vrací počet položek, u kterých došlo k změně.
    """
    if tag_path is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        tag_path = os.path.join(script_dir, TAG_ZOBRAZENI_FILE)
    if not os.path.exists(tag_path):
        print(f"Soubor {tag_path} neexistuje.")
        return 0
    with open(tag_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    count = 0

    def replace_in_obj(obj):
        nonlocal count
        if isinstance(obj, dict):
            for k, v in list(obj.items()):
                if k == "_popis_kategorii":
                    continue
                if isinstance(v, str) and old_substring in v:
                    obj[k] = v.replace(old_substring, new_substring)
                    count += 1
                else:
                    replace_in_obj(v)
        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                if isinstance(item, str) and old_substring in item:
                    obj[i] = item.replace(old_substring, new_substring)
                    count += 1
                else:
                    replace_in_obj(item)

    replace_in_obj(data)
    with open(tag_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return count


if __name__ == "__main__":
    # Získání cesty k adresáři, kde je tento skript
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)
    
    # Export mapy tagů do konfiguračního souboru (pro úpravu zobrazení)
    if len(sys.argv) > 1 and sys.argv[1] == "--export-tag-map":
        out_path = os.path.join(script_dir, TAG_ZOBRAZENI_FILE)
        with open(out_path, 'w', encoding='utf-8') as f:
            json.dump(DIAKRITIKA_MAP, f, ensure_ascii=False, indent=2)
        print(f"Mapa zobrazení tagů vyexportována do: {out_path}")
        print("Tento soubor můžete editovat – klíč = hodnota v datech (bez diakritiky), hodnota = text na tlačítku.")
        sys.exit(0)

    # Nahradit všude stejné zobrazení jedním přepsaným textem (všechny výskyty daného textu v tag_zobrazeni.json)
    if len(sys.argv) >= 4 and sys.argv[1] == "--replace-display":
        old_val = sys.argv[2]
        new_val = sys.argv[3]
        tag_path = os.path.join(script_dir, TAG_ZOBRAZENI_FILE)
        n = replace_display_value_everywhere(old_val, new_val, tag_path)
        print(f"V {TAG_ZOBRAZENI_FILE} nahrazeno {n} výskytů zobrazení „{old_val}“ za „{new_val}“.")
        sys.exit(0)

    # Nahradit slovo uvnitř všech zobrazení (podřetězec) – např. všechna "cisti" na "čistí" bez ohledu na okolní text
    if len(sys.argv) >= 4 and sys.argv[1] == "--replace-display-substring":
        old_val = sys.argv[2]
        new_val = sys.argv[3]
        tag_path = os.path.join(script_dir, TAG_ZOBRAZENI_FILE)
        n = replace_display_substring_everywhere(old_val, new_val, tag_path)
        print(f"V {TAG_ZOBRAZENI_FILE} upraveno {n} položek: „{old_val}“ -> „{new_val}“ (ve všech zobrazeních).")
        sys.exit(0)
    
    excel_file = "Finální.xlsx"
    template_file = "Hotovo4.html"
    output_file = "output.html"
    
    if not os.path.exists(excel_file):
        print(f"Chyba: Soubor {excel_file} nebyl nalezen!")
        sys.exit(1)
    
    if not os.path.exists(template_file):
        print(f"Chyba: Soubor {template_file} nebyl nalezen!")
        sys.exit(1)
    
    try:
        generate_html(excel_file, template_file, output_file)
    except Exception as e:
        print(f"Chyba při generování HTML: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
