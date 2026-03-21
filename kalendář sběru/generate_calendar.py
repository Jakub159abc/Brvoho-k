#!/usr/bin/env python
# -*- coding: utf-8 -*-
import pandas as pd
import re
import os
import html

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Get current directory
script_dir = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(script_dir, 'Finální.xlsx')

# Sloupce Excelu: B=1, C=2, D=3, E=4, F=5, ..., L=11 (sber)
COL_B_INDEX = 1   # český název (s diakritikou)
COL_C_INDEX = 2   # latinský název (s diakritikou)
COL_URL_INDEX = 5   # F = url
COL_SBER_INDEX = 11  # L = sber (část:měsíc | měsíc; ...)

def safe_cell_str(val):
    """Převede buňku na řetězec pro zobrazení; zachová diakritiku."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    if hasattr(val, 'strftime'):  # datetime
        return val.strftime('%d.%m.%Y')
    s = str(val).strip()
    if s == '' or s.lower() == 'nan':
        return ''
    return s


def safe_html(s):
    """Pro zobrazení v HTML: escapuje <>& a zachová diakritiku (UTF-8)."""
    if not s:
        return ''
    return html.escape(s, quote=True)

# Načtení Excelu bez interpretace první řádky jako hlavičky – sloupce podle pozice (0=A, 1=B, 2=C, …)
print(f"Reading Excel file: {excel_path}")
df_raw = pd.read_excel(excel_path, header=None)
if df_raw.shape[1] < 3:
    raise SystemExit(f"Excel musí mít alespoň sloupce A, B, C. Nalezeno sloupců: {df_raw.shape[1]}")

# Rozpoznat, zda první řádek je hlavička (pak ho přeskočíme), nebo první záznam
def looks_like_header_row(row_series):
    if len(row_series) < 3:
        return False
    b_val = safe_cell_str(row_series.iloc[1])
    c_val = safe_cell_str(row_series.iloc[2])
    header_words = ('název', 'nazev', 'český', 'cesky', 'latinský', 'latinsky', 'czech', 'latin', 'rostlina')
    b_lower, c_lower = b_val.lower(), c_val.lower()
    if not b_val or not c_val:
        return True  # prázdná hlavička
    if any(w in b_lower or w in c_lower for w in header_words):
        return True
    if len(b_val) <= 3 and len(c_val) <= 4:
        return True  # např. "B" / "C" nebo zkratky
    return False

first_row = df_raw.iloc[0]
header_skipped = looks_like_header_row(first_row) and len(df_raw) > 1
if header_skipped:
    df = df_raw.iloc[1:].reset_index(drop=True)
    print(f"Načteno {len(df)} řádků dat (první řádek = hlavička, přeskočen). Sloupce B a C = názvy.")
else:
    df = df_raw.reset_index(drop=True)
    print(f"Načteno {len(df)} řádků dat. Sloupce B a C = názvy.")

# Názvy načítáme přímo z Excelu přes openpyxl (bez pandas), aby diakritika zůstala
wb_sheet = None
wb = None
if HAS_OPENPYXL:
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        wb_sheet = wb.active
        print("Názvy (B, C) se berou přímo z Excelu (openpyxl) – diakritika zachována.")
    except Exception as e:
        print(f"Poznámka: openpyxl nelze použít ({e}), názvy z pandas.")
else:
    print("Poznámka: openpyxl není nainstalován (pip install openpyxl), názvy z pandas.")

# Mapování měsíců (české názvy)
MONTHS = ['Leden', 'Únor', 'Březen', 'Duben', 'Květen', 'Červen', 
          'Červenec', 'Srpen', 'Září', 'Říjen', 'Listopad', 'Prosinec']
MONTHS_LOWER = [m.lower() for m in MONTHS]

# Mapování částí rostlin na emoji (každá část má unikátní emoji)
PART_EMOJIS = {
    'květ': '🌸',
    'květy': '💐',
    'květenství': '🌺',
    'list': '🍃',
    'listy': '🍃',
    'listí': '🍃',
    'plod': '🍎',
    'plody': '🍎',
    'kořen': '🌿',
    'kořeny': '🌿',
    'oddenek': '🥔',
    'oddenky': '🥔',
    'nať': '🌱',
    'bylina': '🌿',
    'semena': '🌰',
    'semeno': '🌰',
    'stonky': '🌾',
    'stonek': '🌾',
    'pupeny': '🍀',
    'pupen': '🍀',
    'květní': '🌼',
    'květný': '🌼',
    'blizna': '🌺',
    'blizny': '🌺',
    'tyčinka': '🌼',
    'tyčinky': '🌼',
    'okvětí': '🌼',
    'okvětní': '🌼',
    'plátek': '🌼',
    'plátky': '🌼',
    'květena': '🌺',
    'květeny': '🌺',
    'poupata': '🍀',
    'koren': '🌿',
    'nat': '🌱',
    'cibule': '🥔',
    'hliza': '🥔',
    'hlíza': '🥔',
    'jehlice': '🌿',
    'cela_rasa': '🌿',
    'kura': '🌾',
    'makovice': '🌰',
    'miza': '🌿',
    'pryskyrice': '🟡',
    'samici_sistice': '🌸',
    'stopka': '🌾',
    'stvol': '🌾',
    'plodnice': '🍄',
}

# Mapování částí rostlin na české názvy pro legendu
PART_NAMES = {
    'květ': 'Květ',
    'květy': 'Květy',
    'květenství': 'Květenství',
    'list': 'List',
    'listy': 'Listy',
    'listí': 'Listí',
    'plod': 'Plod',
    'plody': 'Plody',
    'kořen': 'Kořen',
    'kořeny': 'Kořeny',
    'oddenek': 'Oddenek',
    'oddenky': 'Oddenky',
    'nať': 'Nať',
    'bylina': 'Bylina',
    'semena': 'Semena',
    'semeno': 'Semeno',
    'stonky': 'Stonky',
    'stonek': 'Stonek',
    'pupeny': 'Pupeny',
    'pupen': 'Pupen',
    'květní': 'Květní části',
    'květný': 'Květní části',
    'blizna': 'Blizna',
    'blizny': 'Blizny',
    'tyčinka': 'Tyčinka',
    'tyčinky': 'Tyčinky',
    'okvětí': 'Okvětí',
    'okvětní': 'Okvětní',
    'plátek': 'Plátek',
    'plátky': 'Plátky',
    'květena': 'Květena',
    'květeny': 'Květeny',
    'poupata': 'Poupata',
    'koren': 'Kořen',
    'nat': 'Nať',
    'cibule': 'Cibule',
    'hliza': 'Hlíza',
    'hlíza': 'Hlíza',
    'jehlice': 'Jehlice',
    'cela_rasa': 'Celá řasa',
    'kura': 'Kůra',
    'makovice': 'Makovice',
    'miza': 'Míza',
    'pryskyrice': 'Pryskyřice',
    'samici_sistice': 'Samičí šištice',
    'stopka': 'Stopka',
    'stvol': 'Stvol',
    'plodnice': 'Plodnice',
}

def get_emoji_for_part(part_text):
    """Najde emoji pro část rostliny"""
    if not part_text or pd.isna(part_text):
        return ''
    
    part_lower = str(part_text).lower().strip()
    
    # Zkontroluj všechny klíče
    for key, emoji in PART_EMOJIS.items():
        if key in part_lower:
            return emoji
    
    # Default emoji
    return '🌿'

def remove_diacritics(text):
    """Odstraní diakritiku z textu pro lepší porovnávání"""
    replacements = {
        'á': 'a', 'č': 'c', 'ď': 'd', 'é': 'e', 'ě': 'e', 'í': 'i', 'ň': 'n',
        'ó': 'o', 'ř': 'r', 'š': 's', 'ť': 't', 'ú': 'u', 'ů': 'u', 'ý': 'y', 'ž': 'z',
        'Á': 'A', 'Č': 'C', 'Ď': 'D', 'É': 'E', 'Ě': 'E', 'Í': 'I', 'Ň': 'N',
        'Ó': 'O', 'Ř': 'R', 'Š': 'S', 'Ť': 'T', 'Ú': 'U', 'Ů': 'U', 'Ý': 'Y', 'Ž': 'Z'
    }
    result = text
    for old, new in replacements.items():
        result = result.replace(old, new)
    return result

def parse_collection_data(text):
    """Parsuje text ze sloupce J a vrátí dict s měsíci a částmi"""
    if pd.isna(text) or not text:
        return {}
    
    text = str(text).strip()
    if not text:
        return {}
    
    result = {}
    
    # Normalizuj text - převeď na lowercase a odstran diakritiku pro porovnávání
    text_lower = text.lower()
    text_no_diacritics = remove_diacritics(text_lower)
    
    # Mapování měsíců (s diakritikou i bez) - pro každý měsíc všechny varianty
    month_variants_map = {}
    for month_cz in MONTHS:
        month_lower = month_cz.lower()
        variants = [month_lower, remove_diacritics(month_lower)]
        # Přidej různé tvary
        if month_cz == 'leden':
            variants.extend(['lednu', 'ledna', 'ledn', 'led'])
        elif month_cz == 'únor':
            variants.extend(['únoru', 'února', 'unor', 'unoru', 'unora'])
        elif month_cz == 'březen':
            variants.extend(['březnu', 'března', 'brezen', 'breznu', 'brezna', 'brez'])
        elif month_cz == 'duben':
            variants.extend(['dubnu', 'dubna', 'dub'])
        elif month_cz == 'květen':
            variants.extend(['květnu', 'května', 'kveten', 'kvetnu', 'kvetna', 'kvet'])
        elif month_cz == 'červen':
            variants.extend(['červnu', 'června', 'cerven', 'cervnu', 'cervna', 'cerv'])
        elif month_cz == 'červenec':
            variants.extend(['červenci', 'července', 'cervenec', 'cervenci', 'cervence'])
        elif month_cz == 'srpen':
            variants.extend(['srpnu', 'srpna', 'srp'])
        elif month_cz == 'září':
            variants.extend(['zari'])
        elif month_cz == 'říjen':
            variants.extend(['říjnu', 'října', 'rijen', 'rijnu', 'rijna', 'rij'])
        elif month_cz == 'listopad':
            variants.extend(['listopadu', 'listopada'])
        elif month_cz == 'prosinec':
            variants.extend(['prosince', 'prosinci'])
        month_variants_map[month_cz] = variants
    
    # Rozděl text podle středníků na jednotlivé sekce
    # Formát: "část1:měsíc1 | měsíc2;část2:měsíc3 | měsíc4"
    sections = re.split(r'[;]', text_lower)
    
    for section in sections:
        section = section.strip()
        if not section:
            continue
        
        # Najdi dvojtečku - formát "část:měsíc" nebo "část:měsíc1 | měsíc2"
        colon_pos = section.find(':')
        if colon_pos < 0:
            continue
        
        # Najdi část před dvojtečkou
        part_text = section[:colon_pos].strip()
        months_text = section[colon_pos + 1:].strip()
        
        # Najdi emoji pro část - hledej přesnou shodu nebo začátek slova
        part_emoji = None
        part_text_no_diac = remove_diacritics(part_text)
        part_keys_sorted = sorted(PART_EMOJIS.keys(), key=len, reverse=True)
        
        # Nejprve zkus přesnou shodu (s diakritikou i bez)
        if part_text in PART_EMOJIS:
            part_emoji = PART_EMOJIS[part_text]
        elif part_text_no_diac in PART_EMOJIS:
            part_emoji = PART_EMOJIS[part_text_no_diac]
        else:
            # Pak zkus najít část v textu - zkontroluj všechny varianty
            for part_key in part_keys_sorted:
                part_key_no_diac = remove_diacritics(part_key)
                
                # Přesná shoda (s diakritikou i bez)
                if (part_key == part_text or 
                    part_key_no_diac == part_text_no_diac or
                    part_key == part_text_no_diac or
                    part_key_no_diac == part_text):
                    part_emoji = PART_EMOJIS[part_key]
                    break
                
                # Část je v textu (např. "kvet" v "kveten")
                if (part_key in part_text or 
                    part_key_no_diac in part_text_no_diac or
                    part_key in part_text_no_diac or
                    part_key_no_diac in part_text):
                    part_emoji = PART_EMOJIS[part_key]
                    break
                
                # Text je v části (např. "kveten" obsahuje "kvet")
                if (part_text in part_key or 
                    part_text_no_diac in part_key_no_diac or
                    part_text in part_key_no_diac or
                    part_text_no_diac in part_key):
                    part_emoji = PART_EMOJIS[part_key]
                    break
        
        if not part_emoji:
            # Pokud se nepodařilo najít část, zkus ještě jednou s normalizovaným textem
            # Možná je tam mezera nebo jiný znak
            part_text_clean = re.sub(r'[^\w]', '', part_text)
            part_text_clean_no_diac = remove_diacritics(part_text_clean)
            
            for part_key in part_keys_sorted:
                part_key_no_diac = remove_diacritics(part_key)
                
                if (part_key == part_text_clean or 
                    part_key_no_diac == part_text_clean_no_diac or
                    part_text_clean in part_key or
                    part_text_clean_no_diac in part_key_no_diac or
                    part_key in part_text_clean or
                    part_key_no_diac in part_text_clean_no_diac):
                    part_emoji = PART_EMOJIS[part_key]
                    break
        
        if not part_emoji:
            # Sleduj části, které se nepodařilo najít
            unfound_parts.add(part_text)
            continue
        
        # Rozděl měsíce podle svislé čárky nebo středníku
        # Formát: "měsíc1 | měsíc2" nebo "měsíc1;měsíc2"
        month_parts = re.split(r'[|;]', months_text)
        
        for month_part in month_parts:
            month_part = month_part.strip()
            if not month_part:
                continue
            
            # Najdi měsíc v textu - preferuj delší shody (cervenec před cerven)
            month_part_no_diac = remove_diacritics(month_part)
            found_month = None
            best_match_length = 0
            exact_match = False
            
            # Nejdřív zkontroluj přesné shody
            for month_cz, variants in month_variants_map.items():
                for variant in variants:
                    variant_lower = variant.lower()
                    variant_no_diac = remove_diacritics(variant_lower)
                    
                    # Přesná shoda má nejvyšší prioritu
                    if (variant_lower == month_part or 
                        variant_no_diac == month_part_no_diac):
                        found_month = month_cz
                        best_match_length = 999  # Velké číslo pro přesnou shodu
                        exact_match = True
                        break
                
                if exact_match:
                    break
            
            # Pokud jsme nenašli přesnou shodu, hledej částečné shody (preferuj delší)
            if not exact_match:
                for month_cz, variants in month_variants_map.items():
                    for variant in variants:
                        variant_lower = variant.lower()
                        variant_no_diac = remove_diacritics(variant_lower)
                        
                        # Zkontroluj, jestli je variant obsažen v month_part
                        if variant_lower in month_part or variant_no_diac in month_part_no_diac:
                            match_length = len(variant_lower)
                            if match_length > best_match_length:
                                found_month = month_cz
                                best_match_length = match_length
                        # Nebo jestli je month_part obsažen ve variantu
                        elif month_part in variant_lower or month_part_no_diac in variant_no_diac:
                            match_length = len(month_part)
                            if match_length > best_match_length:
                                found_month = month_cz
                                best_match_length = match_length
            
            # Pokud jsme nenašli shodu, zkus ještě jednou s normalizovaným textem
            if not found_month:
                month_part_clean = re.sub(r'[^\w]', '', month_part)
                month_part_clean_no_diac = remove_diacritics(month_part_clean)
                
                for month_cz, variants in month_variants_map.items():
                    for variant in variants:
                        variant_lower = variant.lower()
                        variant_no_diac = remove_diacritics(variant_lower)
                        variant_clean = re.sub(r'[^\w]', '', variant_lower)
                        variant_clean_no_diac = remove_diacritics(variant_clean)
                        
                        if (variant_clean == month_part_clean or 
                            variant_clean_no_diac == month_part_clean_no_diac or
                            variant_clean in month_part_clean or
                            variant_clean_no_diac in month_part_clean_no_diac or
                            month_part_clean in variant_clean or
                            month_part_clean_no_diac in variant_clean_no_diac):
                            match_length = max(len(variant_clean), len(month_part_clean))
                            if match_length > best_match_length:
                                found_month = month_cz
                                best_match_length = match_length
            
            # "celoročně" = zobraz ve všech měsících
            if not found_month and 'celorocne' in remove_diacritics(month_part.lower()):
                for m in MONTHS:
                    if m not in result:
                        result[m] = []
                    if part_emoji not in result[m]:
                        result[m].append(part_emoji)
                found_month = True  # už jsme přidali, níže nic nedělat
            if found_month is True:
                pass  # už přidáno (celoročně)
            elif found_month:
                if found_month not in result:
                    result[found_month] = []
                if part_emoji not in result[found_month]:
                    result[found_month].append(part_emoji)
    
    return result

# Zpracuj data
plants_data = []
plants_without_data = []

# Sleduj všechny použité emoji pro legendu
used_emojis = set()

# Sleduj části, které se nepodařilo najít
unfound_parts = set()

# Seznam problematických rostlin pro debug
problem_plants = [
    'Bakopa drobnolistá', 'Bacopa monnieri', 'Bazalka posvátná', 'Ocimum tenuiflorum',
    'Bazalka pravá', 'Ocimum basilicum', 'Bažanka vytrvalá', 'Helleborus viridis',
    'Benedikt čubet', 'Cnicus benedictus', 'Bez černý'
]

# Načítání: názvy ze sloupců B a C – pokud je k dispozici openpyxl, přímo z Excelu (diakritika)
for idx, row in df.iterrows():
    if wb_sheet is not None:
        excel_row = (idx + 2) if header_skipped else (idx + 1)  # Excel řádky 1-based
        cz_name = safe_cell_str(wb_sheet.cell(row=excel_row, column=2).value)   # B
        lat_name = safe_cell_str(wb_sheet.cell(row=excel_row, column=3).value)  # C
    else:
        cz_name = safe_cell_str(row.iloc[COL_B_INDEX])
        lat_name = safe_cell_str(row.iloc[COL_C_INDEX])
    url = safe_cell_str(row.iloc[COL_URL_INDEX]) if len(row) > COL_URL_INDEX else ''  # F = url
    
    # Data o sběru jsou ve sloupci L (index 11) = "sber" (formát: část:měsíc | měsíc; část:měsíc...)
    collection_info = ''
    if len(row) > COL_SBER_INDEX and pd.notna(row.iloc[COL_SBER_INDEX]):
        collection_info = row.iloc[COL_SBER_INDEX]
    if not collection_info or not str(collection_info).strip():
        for col_idx in range(8, min(13, len(row))):
            val = row.iloc[col_idx] if pd.notna(row.iloc[col_idx]) else ''
            if val and str(val).strip():
                val_str = str(val).lower()
                if ':' in val_str and (any(m in val_str for m in ['leden', 'únor', 'březen', 'duben', 'květen',
                    'červen', 'srpen', 'září', 'říjen', 'listopad', 'prosinec', 'rijen', 'cerven', 'kveten', 'zari']) or
                   any(p in val_str for p in ['květ', 'list', 'plod', 'kořen', 'nať', 'semena', 'koren', 'nat'])):
                    collection_info = val
                    break
    
    if not cz_name:
        continue
    # Vynechat řádek z hlavičky Excelu (zobrazení_cz / zobrazení_lat)
    cz_lower = str(cz_name).strip().lower()
    if cz_lower in ('zobrazení_cz', 'zobrazeni_cz', 'zobrazení_lat', 'zobrazeni_lat'):
        continue
    
    # Zkontroluj, jestli je to problematická rostlina
    is_problem = any(p.lower() in str(cz_name).lower() or p.lower() in str(lat_name).lower() 
                     for p in problem_plants)
    
    collection_months = parse_collection_data(collection_info)
    
    # Debug pro problematické rostliny (bez emoji kvůli výstupu na Windows konzoli)
    if is_problem:
        print(f"\n[DEBUG] - {cz_name}:")
        print(f"  Collection info: '{collection_info}'")
        print(f"  Parsed months count: {len(collection_months)}")
    
    # Debug: zkontroluj, jestli má rostlina data v Excelu, ale ne v parsovaných datech
    if pd.notna(collection_info) and collection_info and str(collection_info).strip():
        if not collection_months:
            plants_without_data.append((cz_name, str(collection_info).strip()))
            if is_problem:
                print(f"  WARNING: Has data but couldn't parse!")
    
    # Sleduj použité emoji
    for month, emojis in collection_months.items():
        for emoji in emojis:
            used_emojis.add(emoji)
    
    plants_data.append({
        'cz_name': cz_name,
        'lat_name': lat_name,
        'url': url,
        'collection': collection_months
    })

if wb is not None:
    try:
        wb.close()
    except Exception:
        pass

print(f"Processed {len(plants_data)} plants")
if plants_without_data:
    print(f"\nWarning: {len(plants_without_data)} plants have data in Excel but couldn't be parsed:")
    for name, data in plants_without_data[:10]:  # Zobraz prvních 10
        print(f"  - {name}: '{data}'")
    if len(plants_without_data) > 10:
        print(f"  ... and {len(plants_without_data) - 10} more")

if unfound_parts:
    print(f"\nWarning: {len(unfound_parts)} parts couldn't be found in mapping:")
    for part in sorted(unfound_parts):
        print(f"  - '{part}'")
    print(f"\nTip: Add these parts to PART_EMOJIS if they are valid plant parts.")

# Generate HTML
html_content = f"""<!DOCTYPE html>
<html lang="cs">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Kalendář sběru léčivých rostlin</title>
<style>
* {{
  box-sizing: border-box;
  margin: 0;
  padding: 0;
}}

body {{
  font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
  background: linear-gradient(135deg, #f5f7fa 0%, #e8f5e9 100%);
  color: #2d5016;
  line-height: 1.6;
  padding: 20px;
  min-height: 100vh;
}}

.container {{
  max-width: 1800px;
  margin: 0 auto;
  background: #ffffff;
  border-radius: 20px;
  box-shadow: 0 10px 40px rgba(0,0,0,0.1);
  padding: 30px;
  overflow-x: auto;
}}

h1 {{
  color: #2e7d32;
  font-size: 2.5em;
  margin-bottom: 10px;
  text-align: center;
  text-shadow: 2px 2px 4px rgba(0,0,0,0.1);
  font-weight: 700;
}}

.subtitle {{
  text-align: center;
  color: #558b2f;
  margin-bottom: 30px;
  font-size: 1.1em;
}}

.calendar-table {{
  width: 100%;
  border-collapse: collapse;
  margin-top: 20px;
  background: #ffffff;
  border-radius: 15px;
  overflow: hidden;
  box-shadow: 0 3px 10px rgba(0,0,0,0.08);
}}

.calendar-table thead {{
  background: linear-gradient(135deg, #4caf50 0%, #388e3c 100%);
  color: #ffffff;
}}

.calendar-table th {{
  padding: 15px 10px;
  text-align: center;
  font-weight: 600;
  font-size: 0.95em;
  border-right: 1px solid rgba(255,255,255,0.2);
}}

.calendar-table th:first-child {{
  text-align: left;
  padding-left: 20px;
  min-width: 250px;
  background: linear-gradient(135deg, #2e7d32 0%, #1b5e20 100%);
  position: sticky;
  left: 0;
  z-index: 10;
}}

.calendar-table th:last-child {{
  border-right: none;
}}

.calendar-table th.current-month {{
  background: linear-gradient(135deg, #66bb6a 0%, #43a047 100%);
  box-shadow: 0 4px 8px rgba(76, 175, 80, 0.3);
  position: relative;
}}

.calendar-table th.current-month::after {{
  content: '●';
  position: absolute;
  top: 5px;
  right: 5px;
  font-size: 0.7em;
  color: #fff;
}}

.calendar-table tbody tr {{
  border-bottom: 1px solid #e8f5e9;
  transition: background 0.2s ease;
}}

.calendar-table tbody tr:hover {{
  background: #f1f8f4;
}}

.calendar-table td {{
  padding: 12px 10px;
  text-align: center;
  border-right: 1px solid #e8f5e9;
}}

.calendar-table td:first-child {{
  text-align: left;
  padding-left: 20px;
  background: #ffffff;
  position: sticky;
  left: 0;
  z-index: 5;
  font-weight: 500;
  min-width: 250px;
}}

.calendar-table tbody tr:hover td:first-child {{
  background: #f1f8f4;
}}

.calendar-table td:last-child {{
  border-right: none;
}}

.calendar-table td.current-month {{
  background: #e8f5e9;
  border-left: 3px solid #4caf50;
  border-right: 3px solid #4caf50;
  font-weight: 500;
}}

.calendar-table td:first-child.current-month {{
  background: #e8f5e9;
  border-left: 3px solid #4caf50;
  border-right: none;
  z-index: 6;
}}

.calendar-table tbody tr:hover td.current-month {{
  background: #c8e6c9;
}}

.calendar-table tbody tr:hover td:first-child.current-month {{
  background: #c8e6c9;
}}

.calendar-table th.month-cell {{
  cursor: pointer;
  user-select: none;
  transition: background 0.2s ease;
}}

.calendar-table th.selected-month {{
  background: linear-gradient(135deg, #ff9800 0%, #f57c00 100%);
  box-shadow: 0 4px 8px rgba(255, 152, 0, 0.4);
  position: relative;
}}

.calendar-table th.selected-month::after {{
  content: '✓';
  position: absolute;
  top: 5px;
  right: 5px;
  font-size: 0.8em;
  color: #fff;
}}

.calendar-table td.month-cell {{
  cursor: pointer;
  transition: background 0.2s ease;
}}

.calendar-table td.selected-month {{
  background: #fff3e0;
  border-left: 3px solid #ff9800;
  border-right: 3px solid #ff9800;
  font-weight: 500;
}}

.calendar-table td:first-child.selected-month {{
  background: #fff3e0;
  border-left: 3px solid #ff9800;
  border-right: none;
  z-index: 6;
}}

.calendar-table tbody tr:hover td.selected-month {{
  background: #ffe0b2;
}}

.calendar-table tbody tr:hover td:first-child.selected-month {{
  background: #ffe0b2;
}}

.calendar-table th.current-month.selected-month {{
  background: linear-gradient(135deg, #66bb6a 0%, #43a047 100%) !important;
  box-shadow: 0 4px 8px rgba(76, 175, 80, 0.3);
}}
.calendar-table td.current-month.selected-month {{
  background: #e8f5e9 !important;
  border-left: 3px solid #4caf50;
  border-right: 3px solid #4caf50;
}}
.calendar-table td:first-child.current-month.selected-month {{
  background: #e8f5e9 !important;
  border-left: 3px solid #4caf50;
  border-right: none;
}}
.calendar-table tbody tr:hover td.current-month.selected-month {{
  background: #c8e6c9 !important;
}}
.calendar-table tbody tr:hover td:first-child.current-month.selected-month {{
  background: #c8e6c9 !important;
}}

.plant-name {{
  font-weight: 700;
  color: #1b5e20;
  font-size: 1em;
  margin-bottom: 4px;
}}

.plant-cell-link {{
  display: block;
  text-decoration: none;
  color: inherit;
  transition: color 0.2s ease;
}}

.plant-cell-link:hover .plant-name {{
  color: #4caf50;
  text-decoration: underline;
}}

.plant-link {{
  color: #1b5e20;
  text-decoration: none;
  transition: color 0.2s ease;
}}

.plant-link:hover {{
  color: #4caf50;
  text-decoration: underline;
}}

.plant-name-lat {{
  font-size: 0.85em;
  color: #558b2f;
  font-style: italic;
  margin-top: 2px;
}}

.month-cell {{
  min-width: 80px;
}}

.emoji-container {{
  display: flex;
  flex-wrap: wrap;
  gap: 4px;
  justify-content: center;
  align-items: center;
  font-size: 1.2em;
}}

.emoji {{
  font-size: 1.3em;
  cursor: pointer;
  position: relative;
  transition: transform 0.2s ease;
}}

.emoji:hover {{
  transform: scale(1.2);
}}

.emoji-tooltip {{
  position: fixed;
  background: #2e7d32;
  color: #ffffff;
  padding: 8px 12px;
  border-radius: 8px;
  font-size: 0.9em;
  pointer-events: none;
  z-index: 1000;
  box-shadow: 0 4px 12px rgba(0,0,0,0.3);
  opacity: 0;
  transition: opacity 0.2s ease;
  white-space: nowrap;
}}

.emoji-tooltip.show {{
  opacity: 1;
}}

@media (max-width: 768px) {{
  .container {{
    padding: 15px;
  }}
  
  h1 {{
    font-size: 1.8em;
  }}
  
  .calendar-table {{
    font-size: 0.85em;
  }}
  
  .calendar-table th,
  .calendar-table td {{
    padding: 8px 5px;
  }}
  
  .calendar-table th:first-child,
  .calendar-table td:first-child {{
    min-width: 150px;
    padding-left: 10px;
  }}
  
  .emoji {{
    font-size: 1.1em;
  }}
}}

.legend {{
  margin-bottom: 30px;
  padding: 20px;
  background: #f1f8f4;
  border-radius: 15px;
  border: 2px solid #c8e6c9;
  box-shadow: 0 2px 8px rgba(0,0,0,0.08);
}}

.legend h2 {{
  color: #2e7d32;
  margin-bottom: 15px;
  font-size: 1.5em;
}}

.legend-items {{
  display: flex;
  flex-wrap: wrap;
  gap: 20px;
}}

.legend-item {{
  display: flex;
  align-items: center;
  gap: 8px;
  font-size: 1.1em;
}}

.legend-emoji {{
  font-size: 1.5em;
}}

.intro {{
  max-width: 720px;
  margin: 0 auto 24px;
  padding: 0 8px;
  color: #2d5016;
  font-size: 1rem;
  line-height: 1.65;
  text-align: center;
}}
.intro p {{ margin-bottom: 0.75em; }}
.intro p:last-child {{ margin-bottom: 0; }}
.intro strong {{ color: #1b5e20; }}

.how-to-use {{
  max-width: 640px;
  margin: 0 auto 20px;
  padding: 0 8px;
}}
.how-to-use details {{
  overflow: hidden;
}}
.how-to-use summary {{
  text-align: center;
  padding: 8px 0;
  font-weight: 700;
  color: #1b5e20;
  cursor: pointer;
  list-style: none;
  user-select: none;
  background: none;
  border: none;
  outline: none;
}}
.how-to-use summary::-webkit-details-marker {{ display: none; }}
.how-to-use summary::marker {{ display: none; }}
.how-to-use summary::before {{
  content: '▶ ';
  font-size: 0.75em;
  color: #558b2f;
  margin-right: 4px;
}}
.how-to-use details[open] summary::before {{
  content: '▼ ';
}}
.how-to-use .how-to-use-content {{
  padding: 0 16px 16px;
  color: #2d5016;
  font-size: 1rem;
  line-height: 1.65;
}}
.how-to-use .how-to-use-content p {{ margin: 0; }}

.search-wrap {{
  margin-bottom: 20px;
  max-width: 400px;
  margin-left: auto;
  margin-right: auto;
  position: relative;
}}
.search-input {{
  width: 100%;
  padding: 12px 16px;
  font-size: 1rem;
  border: 2px solid #c8e6c9;
  border-radius: 12px;
  outline: none;
  transition: border-color 0.2s;
}}
.search-input:focus {{
  border-color: #4caf50;
}}
.search-dropdown {{
  position: absolute;
  top: 100%;
  left: 0;
  right: 0;
  margin-top: 4px;
  background: #fff;
  border: 2px solid #c8e6c9;
  border-radius: 12px;
  max-height: 280px;
  overflow-y: auto;
  box-shadow: 0 8px 24px rgba(0,0,0,0.12);
  z-index: 20;
  display: none;
}}
.search-dropdown.show {{ display: block; }}
.search-dropdown-item {{
  padding: 10px 16px;
  cursor: pointer;
  border-bottom: 1px solid #e8f5e9;
  display: block;
  text-decoration: none;
  color: #1b5e20;
}}
.search-dropdown-item:hover {{
  background: #e8f5e9;
}}
.search-dropdown-item .plant-lat {{ font-size: 0.85em; color: #558b2f; font-style: italic; margin-top: 2px; }}

.month-summary {{
  margin-bottom: 20px;
  padding: 20px;
  background: #fff3e0;
  border: 2px solid #ffb74d;
  border-radius: 15px;
  display: none;
}}
.month-summary.show {{ display: block; }}
.month-summary h2 {{
  color: #e65100;
  font-size: 1.25rem;
  margin-bottom: 12px;
}}
.month-summary-list {{
  list-style: none;
  columns: 1;
  column-gap: 24px;
}}
@media (min-width: 600px) {{ .month-summary-list {{ columns: 2; }} }}
@media (min-width: 900px) {{ .month-summary-list {{ columns: 3; }} }}
.month-summary-list li {{
  break-inside: avoid;
  margin-bottom: 8px;
  padding: 6px 0;
  border-bottom: 1px solid #ffe0b2;
}}
.month-summary-list .plant-name-inline {{ font-weight: 600; color: #1b5e20; }}
.month-summary-list .parts-inline {{ font-size: 0.9em; color: #558b2f; }}
</style>
</head>
<body>
<div class="container">
  <h1>Kalendář sběru léčivých rostlin</h1>
  <p class="subtitle">Tato tabulka slouží jako přehled vhodných termínů sběru jednotlivých částí léčivých rostlin v průběhu roku.</p>
  
  <div class="how-to-use">
    <details>
      <summary>Jak tabulku používat:</summary>
      <div class="how-to-use-content">
        <p>Tabulkou můžeš volně listovat a podle znaků nacházet sbírané části rostlin v jednotlivých měsících. Aktuální měsíc je označen tmavší barvou. V poli „Hledat rostlinu“ můžete vyhledat konkrétní bylinku podle českého názvu. Kliknutím na název měsíce v hlavičce tabulky nebo na libovolnou buňku v daném sloupci měsíc označíte a nad tabulkou se zobrazí seznam všech bylinek a jejich částí, které se v daný měsíc dají sbírat. Dalším kliknutím na stejný měsíc označení zrušíte. Odkazy u názvů rostlin vedou na podrobné informace o této rostlině, které postupně připravujeme.</p>
      </div>
    </details>
  </div>
  
  <div class="search-wrap">
    <input type="text" class="search-input" id="plantSearch" placeholder="Hledat rostlinu" autocomplete="off" aria-label="Vyhledat rostlinu">
    <div class="search-dropdown" id="searchDropdown" role="listbox" aria-label="Nalezené rostliny"></div>
  </div>
  
  <div class="month-summary" id="monthSummary">
    <h2 id="monthSummaryTitle">Rostliny pro vybraný měsíc</h2>
    <ul class="month-summary-list" id="monthSummaryList"></ul>
  </div>
  
  <div class="legend">
    <h2>Legenda</h2>
    <div class="legend-items">
"""

# Generate legend HTML - zobrazíme pouze použité emoji
# Vytvoř mapování emoji -> nejlepší název části
emoji_to_names = {}
for part_key, emoji in PART_EMOJIS.items():
    if part_key in PART_NAMES:
        part_name = PART_NAMES[part_key]
        if emoji not in emoji_to_names:
            emoji_to_names[emoji] = []
        emoji_to_names[emoji].append((part_key, part_name))

# Pro každé použité emoji najdi nejlepší název a zobraz ho
# Vytvoř mapování emoji -> název pro použití v tooltipech
emoji_name_map = {}
legend_items = []
for emoji in sorted(used_emojis):
    if emoji in emoji_to_names:
        # Vezmi první název (nebo můžeme použít nejdelší/nejkratší)
        # Použijeme první, protože PART_NAMES je už seřazené podle důležitosti
        part_key, part_name = emoji_to_names[emoji][0]
        emoji_name_map[emoji] = part_name
        legend_items.append((emoji, part_name))

# Seřaď podle názvu části
legend_items.sort(key=lambda x: x[1])

# Vygeneruj HTML pro legendu
for emoji, part_name in legend_items:
    html_content += f'      <div class="legend-item">\n'
    html_content += f'        <span class="legend-emoji">{emoji}</span>\n'
    html_content += f'        <span>{part_name}</span>\n'
    html_content += f'      </div>\n'

html_content += """    </div>
  </div>
  
  <table class="calendar-table">
    <thead>
      <tr>
        <th>Rostlina</th>
"""

# Add month headers
for month in MONTHS:
    html_content += f'        <th class="month-cell">{month}</th>\n'

html_content += """      </tr>
    </thead>
    <tbody>
"""

# Add plant rows (data-plant-name pro vyhledávání)
for plant in plants_data:
    cz_display = safe_html(plant["cz_name"])
    lat_display = safe_html(plant["lat_name"])
    html_content += f'      <tr data-plant-name="{cz_display}">\n'
    html_content += '        <td>\n'
    url_attr = safe_html(plant["url"]) if plant["url"] else ""
    if plant['url']:
        html_content += f'          <a href="{url_attr}" target="_blank" class="plant-cell-link">\n'
        html_content += f'            <div class="plant-name">{cz_display}</div>\n'
        if plant['lat_name']:
            html_content += f'            <div class="plant-name-lat">{lat_display}</div>\n'
        html_content += '          </a>\n'
    else:
        html_content += f'          <div class="plant-name">{cz_display}</div>\n'
        if plant['lat_name']:
            html_content += f'          <div class="plant-name-lat">{lat_display}</div>\n'
    html_content += '        </td>\n'
    
    # Add month cells
    for month in MONTHS:
        html_content += '        <td class="month-cell">\n'
        if month in plant['collection']:
            emojis = plant['collection'][month]
            html_content += '          <div class="emoji-container">\n'
            for emoji in emojis:
                # Získej název části pro tooltip
                part_name = emoji_name_map.get(emoji, 'Část rostliny')
                html_content += f'            <span class="emoji" data-part-name="{part_name}">{emoji}</span>\n'
            html_content += '          </div>\n'
        html_content += '        </td>\n'
    
    html_content += '      </tr>\n'

html_content += f"""    </tbody>
  </table>
  
</div>

<script>
document.addEventListener("DOMContentLoaded", function() {{
  // Mapování měsíců: JavaScript měsíc (0-11) -> český název
  const monthNames = [
    "Leden", "Únor", "Březen", "Duben", "Květen", "Červen",
    "Červenec", "Srpen", "Září", "Říjen", "Listopad", "Prosinec"
  ];
  
  const currentDate = new Date();
  const currentMonthIndex = currentDate.getMonth();
  const currentMonthName = monthNames[currentMonthIndex];
  
  const table = document.querySelector('.calendar-table');
  const thElements = table ? table.querySelectorAll('thead tr:first-child th') : document.querySelectorAll('.calendar-table thead th');
  const allRows = table ? table.querySelectorAll('tbody tr') : document.querySelectorAll('.calendar-table tbody tr');
  
  function highlightCurrentMonthColumn() {{
    if (!thElements.length || !allRows.length) return;
    var colIndex = -1;
    for (var i = 0; i < thElements.length; i++) {{
      if (thElements[i].textContent.trim() === currentMonthName) {{ colIndex = i; break; }}
    }}
    if (colIndex < 0) return;
    for (var i = 0; i < thElements.length; i++) {{
      if (i === colIndex) thElements[i].classList.add('current-month');
      else thElements[i].classList.remove('current-month');
    }}
    for (var r = 0; r < allRows.length; r++) {{
      var tds = allRows[r].querySelectorAll('td');
      for (var c = 0; c < tds.length; c++) {{
        if (c === colIndex) tds[c].classList.add('current-month');
        else tds[c].classList.remove('current-month');
      }}
    }}
  }}
  highlightCurrentMonthColumn();
  setTimeout(highlightCurrentMonthColumn, 100);
  
  const monthSummaryEl = document.getElementById('monthSummary');
  const monthSummaryTitleEl = document.getElementById('monthSummaryTitle');
  const monthSummaryListEl = document.getElementById('monthSummaryList');
  let currentSelectedColumn = -1;
  
  function updateMonthSummary(columnIndex) {{
    if (columnIndex < 1) {{ monthSummaryEl.classList.remove('show'); return; }}
    const monthName = thElements[columnIndex].textContent.trim();
    monthSummaryTitleEl.textContent = 'Rostliny pro ' + monthName;
    monthSummaryListEl.innerHTML = '';
    const items = [];
    allRows.forEach(row => {{
      const tds = row.querySelectorAll('td');
      const td = tds[columnIndex];
      if (!td) return;
      const emojis = td.querySelectorAll('.emoji');
      if (emojis.length === 0) return;
      const plantName = (row.querySelector('.plant-name') || {{}}).textContent.trim() || '';
      const seen = {{}};
      const partsWithEmoji = [];
      Array.from(emojis).forEach(e => {{
        const part = e.getAttribute('data-part-name') || '';
        if (!part || seen[part]) return;
        seen[part] = true;
        const emoji = (e.textContent || '').trim();
        partsWithEmoji.push(emoji ? emoji + ' ' + part : part);
      }});
      const partsStr = partsWithEmoji.join(', ');
      if (plantName) items.push({{ name: plantName, parts: partsStr }});
    }});
    items.forEach(item => {{
      const li = document.createElement('li');
      li.innerHTML = '<span class="plant-name-inline">' + escapeHtml(item.name) + '</span>' +
        (item.parts ? ' <span class="parts-inline">(' + item.parts + ')</span>' : '');
      monthSummaryListEl.appendChild(li);
    }});
    monthSummaryEl.classList.add('show');
  }}
  function escapeHtml(s) {{
    const div = document.createElement('div');
    div.textContent = s;
    return div.innerHTML;
  }}
  
  function highlightColumn(columnIndex) {{
    const alreadySelected = currentSelectedColumn === columnIndex && columnIndex > 0;
    if (alreadySelected) {{
      currentSelectedColumn = -1;
      thElements.forEach((th, idx) => {{
        if (idx > 0) th.classList.remove('selected-month');
      }});
      allRows.forEach(row => {{
        row.querySelectorAll('td').forEach(td => td.classList.remove('selected-month'));
      }});
      monthSummaryEl.classList.remove('show');
      return;
    }}
    currentSelectedColumn = columnIndex;
    thElements.forEach((th, idx) => {{
      if (idx > 0 && idx === columnIndex) th.classList.add('selected-month');
      else if (idx > 0) th.classList.remove('selected-month');
    }});
    allRows.forEach(row => {{
      const tds = row.querySelectorAll('td');
      tds.forEach((td, idx) => {{
        if (idx === columnIndex) td.classList.add('selected-month');
        else td.classList.remove('selected-month');
      }});
    }});
    if (columnIndex > 0) updateMonthSummary(columnIndex);
    else monthSummaryEl.classList.remove('show');
    highlightCurrentMonthColumn();
  }}
  
  thElements.forEach((th, columnIndex) => {{
    if (columnIndex > 0) {{
      th.addEventListener('click', function() {{ highlightColumn(columnIndex); }});
    }}
  }});
  allRows.forEach(row => {{
    const tds = row.querySelectorAll('td.month-cell');
    tds.forEach((td, cellIndex) => {{
      const columnIndex = cellIndex + 1;
      td.addEventListener('click', function() {{ highlightColumn(columnIndex); }});
    }});
  }});
  
  // Tooltip pro emoji
  const tooltip = document.createElement('div');
  tooltip.className = 'emoji-tooltip';
  document.body.appendChild(tooltip);
  
  // Najdi všechny emoji elementy
  const emojiElements = document.querySelectorAll('.emoji');
  
  emojiElements.forEach(emoji => {{
    emoji.addEventListener('click', function(e) {{
      e.stopPropagation(); // Zabraň propagaci kliknutí na buňku
      const partName = this.getAttribute('data-part-name');
      if (partName) {{
        tooltip.textContent = partName;
        tooltip.classList.add('show');
        
        // Pozicuj tooltip vedle kliknutého emoji
        const rect = this.getBoundingClientRect();
        tooltip.style.left = (rect.left + rect.width / 2 - tooltip.offsetWidth / 2) + 'px';
        tooltip.style.top = (rect.top - tooltip.offsetHeight - 10) + 'px';
        
        // Skryj tooltip po 2 sekundách
        setTimeout(() => {{
          tooltip.classList.remove('show');
        }}, 2000);
      }}
    }});
    
    // Skryj tooltip při kliknutí mimo
    emoji.addEventListener('mouseleave', function() {{
      setTimeout(() => {{
        if (!tooltip.matches(':hover')) {{
          tooltip.classList.remove('show');
        }}
      }}, 100);
    }});
  }});
  
  // Skryj tooltip při kliknutí kdekoli jinde
  document.addEventListener('click', function(e) {{
    if (!e.target.classList.contains('emoji')) {{
      tooltip.classList.remove('show');
    }}
  }});
}});
</script>

<script>
(function() {{
  function initPlantSearch() {{
    var input = document.getElementById('plantSearch');
    var dropdown = document.getElementById('searchDropdown');
    if (!input || !dropdown) return;
    function norm(t) {{
      if (t == null) return '';
      t = String(t).toLowerCase();
      try {{ t = t.normalize('NFD').replace(/[\\u0300-\\u036f]/g, ''); }} catch (e) {{}}
      return t;
    }}
    function matchFirstThreeWords(name, normalizedQuery) {{
      var words = String(name || '').trim().split(/\\s+/).slice(0, 3);
      for (var w = 0; w < words.length; w++) {{
        if (norm(words[w]).indexOf(normalizedQuery) === 0) return true;
      }}
      return false;
    }}
    function runSearch() {{
      var q = (input.value || '').trim();
      dropdown.classList.remove('show');
      dropdown.innerHTML = '';
      if (q.length < 1) return;
      var nq = norm(q);
      var rows = document.querySelectorAll('.calendar-table tbody tr');
      var list = [];
      for (var i = 0; i < rows.length; i++) {{
        var row = rows[i];
        var name = (row.getAttribute('data-plant-name') || '').trim();
        if (!name) {{
          var el = row.querySelector('.plant-name');
          name = (el ? el.textContent : '').trim();
        }}
        if (!name || !matchFirstThreeWords(name, nq)) continue;
        var latEl = row.querySelector('.plant-name-lat');
        list.push({{ row: row, name: name, lat: (latEl ? latEl.textContent : '').trim() }});
      }}
      for (var j = 0; j < list.length && j < 50; j++) {{
        var it = list[j];
        var div = document.createElement('div');
        div.className = 'search-dropdown-item';
        div.setAttribute('role', 'option');
        div.textContent = it.name;
        if (it.lat) {{
          var latDiv = document.createElement('div');
          latDiv.className = 'plant-lat';
          latDiv.textContent = it.lat;
          div.appendChild(latDiv);
        }}
        (function(r) {{
          div.addEventListener('click', function() {{
            r.scrollIntoView({{ behavior: 'smooth', block: 'center' }});
            dropdown.classList.remove('show');
            input.value = '';
          }});
        }})(it.row);
        dropdown.appendChild(div);
      }}
      if (list.length > 0) dropdown.classList.add('show');
    }}
    input.addEventListener('input', runSearch);
    input.addEventListener('focus', function() {{
      if ((input.value || '').trim().length >= 1) runSearch();
    }});
    input.addEventListener('keydown', function(e) {{
      if (e.key === 'Escape') dropdown.classList.remove('show');
    }});
    document.addEventListener('click', function(e) {{
      if (!input.contains(e.target) && !dropdown.contains(e.target)) dropdown.classList.remove('show');
    }});
  }}
  if (document.readyState === 'loading') {{
    document.addEventListener('DOMContentLoaded', initPlantSearch);
  }} else {{
    initPlantSearch();
  }}
}})();
</script>

</body>
</html>
"""

# Save HTML file
output_path = os.path.join(script_dir, 'kalendar_sberu.html')
with open(output_path, 'w', encoding='utf-8') as f:
    f.write(html_content)

print(f"HTML calendar generated: {output_path}")
