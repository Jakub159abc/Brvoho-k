# -*- coding: utf-8 -*-
import sys
import os

# Add current directory to path
script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, script_dir)

try:
    import pandas as pd
    import openpyxl
    
    excel_path = os.path.join(script_dir, 'Finální.xlsx')
    print(f"Trying to read: {excel_path}")
    print(f"File exists: {os.path.exists(excel_path)}")
    
    # Try with openpyxl first
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active
    
    print(f"Max row: {ws.max_row}")
    print(f"\nFirst 5 rows of data:")
    for i in range(2, min(7, ws.max_row + 1)):
        col_b = ws.cell(i, 2).value  # Column B
        col_c = ws.cell(i, 3).value  # Column C  
        col_j = ws.cell(i, 10).value  # Column J
        print(f"Row {i}:")
        print(f"  B (Czech): {col_b}")
        print(f"  C (Latin): {col_c}")
        print(f"  J (Collection): {col_j}")
        print()
        
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
