#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Parser textu z GPT chatu do tabulky bylinek (CSV/XLSX).
Formát vstupu: (název_tagu)- hodnota
Použití: python parse_gpt_to_table.py [soubor.txt]
         nebo: echo "..." | python parse_gpt_to_table.py
"""

import csv
import re
import sys
from pathlib import Path

# Mapování názvů tagů z GPT na názvy sloupců v CSV
TAG_TO_COLUMN = {
    "nazev_cz": "nazev_cz",
    "nazev_lat": "nazev_lat",
    "skupina": "skupina",
    "nemoci": "nemoci",
    "tcm_organy": "tcm_organy",
    "ucinky": "ucinky",
    "Cast-rostliny": "Cast-rostliny",
    "sber": "sber",
    "stanoviste": "stanoviste",
    "barva kvetu": "barva kvetu",
    "Bezpečnost": "Bezpečnost",
}

CSV_PATH = Path(__file__).parent / "Finální_upravený.csv"
XLSX_PATH = Path(__file__).parent / "Finální_upravený.xlsx"
URL_BASE = "https://www.brvohorik.cz/l/"


def slug_from_nazev_cz(text: str) -> str:
    """Odvodí URL slug z nazev_cz: lowercase, podtržítka -> pomlčky."""
    return text.strip().lower().replace("_", "-")


def parse_sber_line(line: str) -> str:
    """
    Parsuje řádek (sber) z GPT formátu do formátu tabulky.
    GPT: (sber)-nat: cerven | cervenec (semeno)-semeno: srpen | zari
    Tabulka: nat: cerven | cervenec ; semeno: srpen | zari
    """
    rest = re.sub(r"^\(sber\)\s*-\s*", "", line, flags=re.IGNORECASE).strip()
    parts = []
    # Rozdělit podle ")-" - každý kus pak obsahuje "část: měsíce"
    for segment in re.split(r"\)\s*-\s*", rest):
        segment = segment.strip()
        # Odstranit úvodní "(název" a koncové " (další_část"
        segment = re.sub(r"^\(\w+\)?", "", segment).strip()
        segment = re.sub(r"\s*\(\w+\)?\s*$", "", segment).strip()
        m = re.match(r"^(\w+):\s*(.+)$", segment)
        if m:
            part_name, months = m.group(1), m.group(2).strip()
            parts.append(f"{part_name}: {months}")
    if not parts:
        if rest:
            parts = [s.strip() for s in re.split(r"\s*;\s*", rest)]
    return " ; ".join(p.strip() for p in parts if p.strip())


def parse_gpt_text(text: str) -> dict:
    """
    Parsuje text z GPT a vrátí slovník sloupec -> hodnota.
    """
    result = {}
    lines = text.splitlines()

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Řádek (sber) - speciální zpracování
        if re.match(r"^\(sber\)\s*-", line, re.IGNORECASE):
            result["sber"] = parse_sber_line(line)
            continue

        # Obecný formát (název_tagu)- hodnota (tag může obsahovat mezeru nebo pomlčku)
        m = re.match(r"^\(([^)]+)\)\s*-\s*(.*)$", line)
        if m:
            tag = m.group(1).strip()
            value = m.group(2).strip()
            if tag in TAG_TO_COLUMN:
                result[TAG_TO_COLUMN[tag]] = value

    # URL z nazev_cz
    if "nazev_cz" in result:
        result["url"] = URL_BASE + slug_from_nazev_cz(result["nazev_cz"]) + "/"

    # Výchozí zobrazení pro nové záznamy
    if "zobrazení" not in result:
        result["zobrazení"] = "on"

    return result


def load_csv(path: Path) -> tuple[list[str], list[dict]]:
    """Načte CSV, vrátí (hlavička, seznam řádků jako dict)."""
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        header = reader.fieldnames
        rows = list(reader)
    return header, rows


def save_csv(path: Path, header: list[str], rows: list[dict]) -> None:
    """Uloží CSV (všechna pole v uvozovkách, jako původní tabulka)."""
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=header, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        writer.writerows(rows)


def _normalize_name(s: str) -> str:
    """Normalizace pro porovnání: lowercase, podtržítka -> mezery."""
    return (s or "").strip().lower().replace("_", " ")


def find_row_index(rows: list[dict], data: dict) -> int | None:
    """Najde index řádku podle nazev_lat nebo nazev_cz. Vrátí None pokud nenalezeno."""
    nazev_lat = _normalize_name(data.get("nazev_lat") or "")
    nazev_cz = _normalize_name(data.get("nazev_cz") or "")
    for i, row in enumerate(rows):
        if nazev_lat and _normalize_name(row.get("nazev_lat")) == nazev_lat:
            return i
        if nazev_cz and _normalize_name(row.get("nazev_cz")) == nazev_cz:
            return i
    return None


def next_id(rows: list[dict]) -> str:
    """Vygeneruje další id ve tvaru R00XX."""
    max_num = 0
    for row in rows:
        rid = (row.get("id") or "").strip()
        m = re.match(r"R(\d+)", rid, re.IGNORECASE)
        if m:
            max_num = max(max_num, int(m.group(1)))
    return f"R{max_num + 1:04d}"


def apply_to_table(header: list[str], rows: list[dict], data: dict) -> list[dict]:
    """
    Aktualizuje nebo přidá řádek. data obsahuje pouze sloupce z GPT (včetně url, zobrazení).
    Sloupce, které v data nejsou, se u existujícího řádku nemění; u nového řádku zůstanou prázdné.
    """
    idx = find_row_index(rows, data)
    new_row = {col: "" for col in header}

    if idx is not None:
        # Aktualizace: zkopírovat stávající řádek, pak přepsat jen to, co je v data a není prázdné
        new_row.update(rows[idx])
        for col, val in data.items():
            if col in header and (val or "").strip():
                new_row[col] = val.strip()
        rows[idx] = new_row
    else:
        # Nový řádek
        new_row["id"] = next_id(rows)
        new_row["zobrazení"] = data.get("zobrazení", "on")
        for col, val in data.items():
            if col in header:
                new_row[col] = (val or "").strip()
        rows.append(new_row)

    return rows


def export_xlsx(csv_path: Path, xlsx_path: Path) -> None:
    """Exportuje CSV do XLSX (vyžaduje openpyxl)."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Pro export XLSX nainstalujte: pip install openpyxl", file=sys.stderr)
        return

    header, rows = load_csv(csv_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bylinky"
    for c, name in enumerate(header, 1):
        ws.cell(row=1, column=c, value=name)
    for r, row in enumerate(rows, 2):
        for c, name in enumerate(header, 1):
            ws.cell(row=r, column=c, value=row.get(name, ""))
    wb.save(xlsx_path)
    print(f"XLSX uloženo: {xlsx_path}")


def main() -> None:
    if len(sys.argv) > 1:
        input_path = Path(sys.argv[1])
        if not input_path.exists():
            print(f"Soubor nenalezen: {input_path}", file=sys.stderr)
            sys.exit(1)
        text = input_path.read_text(encoding="utf-8")
    else:
        text = sys.stdin.read()

    if not text.strip():
        print("Žádný vstup.", file=sys.stderr)
        sys.exit(1)

    data = parse_gpt_text(text)
    if not data.get("nazev_lat") and not data.get("nazev_cz"):
        print("Ve vstupu chybí (nazev_lat) nebo (nazev_cz).", file=sys.stderr)
        sys.exit(1)

    if not CSV_PATH.exists():
        print(f"Tabulka nenalezena: {CSV_PATH}", file=sys.stderr)
        sys.exit(1)

    header, rows = load_csv(CSV_PATH)
    idx = find_row_index(rows, data)
    if idx is None:
        nazev = data.get("nazev_cz") or data.get("nazev_lat") or "(neuvedeno)"
        print(
            f"CHYBA: V tabulce nebyl nalezen řádek pro rostlinu „{nazev}“. "
            "Zkontrolujte, že (nazev_cz) nebo (nazev_lat) odpovídá záznamu v tabulce. "
            "Tabulka nebyla změněna.",
            file=sys.stderr,
        )
        sys.exit(1)
    rows = apply_to_table(header, rows, data)
    save_csv(CSV_PATH, header, rows)
    print(f"Tabulka aktualizována: {CSV_PATH}")

    if "--xlsx" in sys.argv or "-x" in sys.argv:
        export_xlsx(CSV_PATH, XLSX_PATH)


if __name__ == "__main__":
    main()
