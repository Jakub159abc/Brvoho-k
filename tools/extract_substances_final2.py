# -*- coding: utf-8 -*-
"""
Z Finální2.xlsx (sloupec „obsahové látky“ / O) vytvoří substance-page-default-from-final2.txt.
Formát buněk: Skupina: látka | látka ; další skupina: …
Spusť z kořene projektu (s nainstalovaným Finální2.xlsx na stejném místě jako dřív):
  python tools/extract_substances_final2.py
"""
import glob
import os
import re
import sys
from collections import OrderedDict

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import load_workbook

SUBGROUP = re.compile(
    r"^([A-Za-zÁČĎÉĚÍŇÓŘŠŤÚŮÝŽáčďéěíňóřštúůýž][^:]{0,55}):\s*(.+)$",
    re.DOTALL,
)
SKIP_ITEM = re.compile(r"^[^:]+:\s*$")
JUNK = frozenset(
    x.lower()
    for x in ("none", "žádné", "není známo", "není známá", "ninguno", "não znamé", "-")
)


def merge(merged, order, g, items):
    g = g.strip()
    if not g:
        return
    if g not in merged:
        merged[g] = set()
        order.append(g)
    for it in items:
        it = it.strip()
        if not it or it.lower() in ("(nenalezeno)", "nenalezeno"):
            continue
        if it.lower() in JUNK:
            continue
        if SKIP_ITEM.match(it):
            continue
        m = SUBGROUP.match(it)
        if m:
            g2, rest = m.group(1).strip(), m.group(2).strip()
            if not rest:
                continue
            merge(merged, order, g2, [x.strip() for x in rest.split("|") if x.strip()])
        else:
            merged[g].add(it)


def parse_cell(text):
    if not text or not str(text).strip():
        return
    t = str(text).replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    for seg in re.split(r"\s*;\s*", t):
        seg = seg.strip()
        if not seg:
            continue
        m = re.match(r"^(.+?):\s*(.+)$", seg, re.DOTALL)
        if not m:
            continue
        yield m.group(1).strip(), [c.strip() for c in m.group(2).split("|") if c.strip()]


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    root = os.path.dirname(here)
    candidates = [
        os.path.join(root, "Finální2.xlsx"),
        os.path.join(root, "Final2.xlsx"),
    ]
    p = next((c for c in candidates if os.path.isfile(c)), None)
    if not p:
        found = glob.glob(os.path.join(root, "Fin*2.xlsx"))
        p = found[0] if found else None
    if not p:
        print("Finální2.xlsx nenalezen vedle editoru.", file=sys.stderr)
        sys.exit(1)
    wb = load_workbook(p, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    hdr = rows[0]
    col_idx = 14
    for i, h in enumerate(hdr):
        if h is None:
            continue
        hs = str(h).lower()
        if "logika" in hs or ("obsahov" in hs and "lát" in hs):
            col_idx = i
            break
    merged = OrderedDict()
    order = []
    for row in rows[1:]:
        if col_idx >= len(row):
            continue
        v = row[col_idx]
        if v is None or not str(v).strip():
            continue
        for g, chunks in parse_cell(v):
            merge(merged, order, g, chunks)
    intro = (
        "Níže je přehled skupin obsahových látek a konkrétních látek podle tabulky Finální2 "
        "(sloupec s obsahovými látkami). U každé skupiny jsou sloučeny všechny látky, které se "
        "v tabulce pod tuto skupinu kdykoli objevily u libovolné rostliny."
    )
    lines = [intro, ""]
    for g in order:
        lines.append("**" + g + "**")
        for it in sorted(merged[g], key=lambda x: (x.lower(), x)):
            lines.append("- " + it)
        lines.append("")
    dest = os.path.join(root, "substance-page-default-from-final2.txt")
    with open(dest, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines).strip())
    print(dest)
    print("skupin:", len(order), "položek celkem:", sum(len(merged[g]) for g in order))


if __name__ == "__main__":
    main()
