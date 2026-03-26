# -*- coding: utf-8 -*-
"""
Skript pro aktualizaci HTML souboru z Excel tabulky.
Načte Excel soubor a vytvoří kompletní HTML soubor s daty vloženými přímo.
"""
import openpyxl
import json
import os
import sys

def update_html_from_excel(excel_file='duchovni-priciny-nemoci.xlsx', 
                           output_html='index.html',
                           script_js='script.js',
                           style_css='style.css'):
    """
    Načte Excel soubor a vytvoří/aktualizuje HTML soubor s daty vloženými přímo.
    
    Args:
        excel_file: Cesta k Excel souboru
        output_html: Cesta k výstupnímu HTML souboru
        script_js: Cesta k JavaScript souboru s logikou
        style_css: Cesta k CSS souboru
    """
    # Získat adresář, kde je tento skript
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, excel_file)
    output_path = os.path.join(script_dir, output_html)
    script_js_path = os.path.join(script_dir, script_js)
    style_css_path = os.path.join(script_dir, style_css)
    
    print(f"Načítám Excel soubor: {excel_path}")
    
    if not os.path.exists(excel_path):
        print(f"CHYBA: Excel soubor nenalezen: {excel_path}")
        return False
    
    try:
        # Načíst Excel soubor
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Načíst hlavičky z prvního řádku
        headers = [cell.value for cell in ws[1]]
        print(f"Nalezeny sloupce: {headers}")
        
        # Načíst data
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell for cell in row):  # Přeskočit prázdné řádky
                record = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i]:
                        record[headers[i]] = value if value is not None else ''
                if record:
                    data.append(record)
        
        print(f"Načteno {len(data)} záznamů")
        
        # Převedení dat do JavaScript formátu
        js_data = json.dumps(data, ensure_ascii=False, indent=2)
        js_content = 'const allData = ' + js_data + ';'
        
        # Načíst CSS
        print(f"Načítám CSS soubor: {style_css_path}")
        if os.path.exists(style_css_path):
            with open(style_css_path, 'r', encoding='utf-8') as f:
                css_content = f.read()
        else:
            print(f"VAROVÁNÍ: CSS soubor nenalezen: {style_css_path}")
            css_content = ""
        
        # Načíst JavaScript logiku
        print(f"Načítám JavaScript soubor: {script_js_path}")
        if os.path.exists(script_js_path):
            with open(script_js_path, 'r', encoding='utf-8') as f:
                script_content = f.read()
        else:
            print(f"VAROVÁNÍ: JavaScript soubor nenalezen: {script_js_path}")
            script_content = ""
        
        # Vytvořit HTML obsah
        html_content = f'''<!DOCTYPE html>
<html lang="cs">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Duševní příčiny nemocí - Filtrování</title>
    <style>
{css_content}
    </style>
</head>
<body>
    <div class="container">
        <h1>Duševní příčiny nemocí</h1>
        <p class="subtitle">Vyhledejte problém a zjistěte jeho možnou duševní příčinu</p>

        <div id="search-container">
            <div id="search-wrapper">
                <input type="text" id="search-input" placeholder="🔍 Zadejte název problému nebo nemoci..." autocomplete="off">
            </div>
        </div>

        <div id="problems" class="problems-grid">
            <div class="no-results">Načítám data...</div>
        </div>
    </div>

    <script>
{js_content}
    </script>
    <script>
{script_content}
    </script>
</body>
</html>'''
        
        # Uložit HTML soubor
        print(f"Ukládám HTML soubor: {output_path}")
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"✓ HTML soubor byl úspěšně vytvořen!")
        print(f"  Soubor: {output_path}")
        print(f"  Celkem záznamů: {len(data)}")
        
        return True
        
    except Exception as e:
        print(f"CHYBA při zpracování: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    # Zkontrolovat argumenty příkazové řádky
    excel_file = sys.argv[1] if len(sys.argv) > 1 else 'duchovni-priciny-nemoci.xlsx'
    output_html = sys.argv[2] if len(sys.argv) > 2 else 'index.html'
    
    print("=" * 60)
    print("Aktualizace HTML souboru z Excel tabulky")
    print("=" * 60)
    print()
    
    success = update_html_from_excel(excel_file, output_html)
    
    if success:
        print()
        print("=" * 60)
        print("✓ Aktualizace dokončena úspěšně!")
        print("=" * 60)
        print()
        print("HTML soubor nyní obsahuje všechna data z Excel tabulky.")
        print("Můžete ho otevřít přímo v prohlížeči bez externích souborů.")
        sys.exit(0)
    else:
        print()
        print("=" * 60)
        print("✗ Aktualizace selhala!")
        print("=" * 60)
        sys.exit(1)
