# -*- coding: utf-8 -*-
"""
Skript pro aktualizaci HTML aplikace z Excel souboru.
Načte Excel soubor a přepíše data.js s aktuálními daty.
"""
import openpyxl
import json
import os
import sys

def update_from_excel(excel_file='duchovni-priciny-nemoci.xlsx', output_file='data.js'):
    """
    Načte Excel soubor a vytvoří/aktualizuje data.js
    
    Args:
        excel_file: Cesta k Excel souboru
        output_file: Cesta k výstupnímu JS souboru
    """
    # Získat adresář, kde je tento skript
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, excel_file)
    output_path = os.path.join(script_dir, output_file)
    
    print(f"Načítám Excel soubor: {excel_path}")
    
    if not os.path.exists(excel_path):
        print(f"CHYBA: Excel soubor nenalezen: {excel_path}")
        return False
    
    try:
        # Načíst Excel soubor
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Načíst hlavičky z prvního řádku
        headers = [cell.value for cell in ws[1]]
        print(f"Nalezeny sloupce: {headers}")
        
        # Načíst data
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell for cell in row):  # Přeskočit prázdné řádky
                record = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i]:
                        record[headers[i]] = value if value is not None else ''
                if record:
                    data.append(record)
        
        print(f"Načteno {len(data)} záznamů")
        
        # Uložit jako JavaScript proměnnou
        js_data = json.dumps(data, ensure_ascii=False, indent=2)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('const allData = ' + js_data + ';')
        
        print(f"✓ Data úspěšně uložena do: {output_path}")
        print(f"✓ Celkem záznamů: {len(data)}")
        
        return True
        
    except Exception as e:
        print(f"CHYBA při zpracování: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    # Zkontrolovat argumenty příkazové řádky
    excel_file = sys.argv[1] if len(sys.argv) > 1 else 'duchovni-priciny-nemoci.xlsx'
    output_file = sys.argv[2] if len(sys.argv) > 2 else 'data.js'
    
    print("=" * 60)
    print("Aktualizace HTML aplikace z Excel souboru")
    print("=" * 60)
    print()
    
    success = update_from_excel(excel_file, output_file)
    
    if success:
        print()
        print("=" * 60)
        print("✓ Aktualizace dokončena úspěšně!")
        print("=" * 60)
        sys.exit(0)
    else:
        print()
        print("=" * 60)
        print("✗ Aktualizace selhala!")
        print("=" * 60)
        sys.exit(1)
